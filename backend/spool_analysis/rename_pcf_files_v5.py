import math
import os
import re
from pathlib import Path
from itertools import permutations
from collections import Counter, defaultdict

import chardet
import pandas as pd

import pcf_rule_table as rule_table

ENABLE_EXCEL_FORMATTING = False

BACKEND_DIR = Path(__file__).resolve().parents[1]
DEFAULT_FILE_DIR = BACKEND_DIR / 'file' / 'parser'


# Compatibility alias for older code paths that still import this constant.
COMPONENT_SKIP_TYPES = set(rule_table.NON_MATERIAL_BLOCK_TYPES)


def read_pcf_text(file_path):
    """Read a PCF file using detected encoding first, then safe fallbacks."""
    encodings = ['utf-8-sig', 'utf-8', 'gbk', 'gb2312', 'cp936', 'big5', 'utf-16']
    raw_content = None

    try:
        with open(file_path, 'rb') as file:
            raw_content = file.read()
    except Exception as exc:
        print(f"读取文件失败: {file_path}, 错误: {exc}")
        return None, None

    candidate_encodings = []
    detected_encoding = detect_file_encoding_from_bytes(raw_content)
    trusted_detected_encodings = {'utf-8', 'utf-8-sig', 'utf-16', 'gbk', 'gb2312', 'cp936', 'big5'}
    if detected_encoding in trusted_detected_encodings:
        candidate_encodings.append(detected_encoding)

    if raw_content.startswith(b'\xef\xbb\xbf'):
        candidate_encodings.insert(0, 'utf-8-sig')
    elif raw_content.startswith((b'\xff\xfe', b'\xfe\xff')):
        candidate_encodings.insert(0, 'utf-16')

    for encoding in encodings:
        if encoding not in candidate_encodings:
            candidate_encodings.append(encoding)

    # Single-byte detections such as latin-1 / ISO-8859-1 / IBM866 can decode
    # arbitrary GBK bytes without error and produce mojibake. Try them only
    # after UTF and Chinese encodings have failed.
    if detected_encoding and detected_encoding not in candidate_encodings:
        candidate_encodings.append(detected_encoding)

    for encoding in candidate_encodings:
        try:
            text = raw_content.decode(encoding)
            print(f"成功使用 {encoding} 编码读取文件 {file_path}")
            return text, encoding
        except UnicodeDecodeError:
            continue
        except Exception as exc:
            print(f"尝试使用 {encoding} 读取 {file_path} 时出错: {exc}")

    try:
        text = raw_content.decode('latin-1')
        print(f"警告: 使用 latin-1 兜底读取文件 {file_path}，可能导致中文乱码")
        return text, 'latin-1'
    except Exception:
        pass

    print(f"无法读取文件 {file_path}，已尝试编码: {', '.join(candidate_encodings)}")
    return None, None


def detect_file_encoding_from_bytes(raw_content):
    """Detect encoding with chardet, normalizing common aliases."""
    if not raw_content:
        return None
    try:
        result = chardet.detect(raw_content)
    except Exception as exc:
        print(f"检测文件编码失败: {exc}")
        return None

    encoding = normalize_detected_encoding(result.get('encoding'))
    confidence = result.get('confidence')
    if encoding:
        confidence_text = f"{confidence:.3f}" if isinstance(confidence, (int, float)) else str(confidence)
        print(f"检测到文件编码: {encoding}, 置信度: {confidence_text}")
    return encoding


def normalize_detected_encoding(encoding):
    if not encoding:
        return None
    normalized = str(encoding).strip().lower()
    alias_map = {
        'utf_8': 'utf-8',
        'utf-8-sig': 'utf-8-sig',
        'utf_16': 'utf-16',
        'gb2312': 'gb2312',
        'gbk': 'gbk',
        'gb18030': 'gbk',
        'cp936': 'cp936',
        'big5': 'big5',
        'ascii': 'utf-8',
        'windows-1252': 'cp1252',
        'iso-8859-1': 'latin-1',
        'latin-1': 'latin-1',
    }
    return alias_map.get(normalized, normalized)


def safe_float_convert(value):
    """Safely convert a string to float."""
    try:
        return float(str(value).strip())
    except (ValueError, TypeError):
        return None


def extract_number(spec):
    """Extract the first numeric part from a spec string."""
    match = re.search(r'\d+(?:\.\d+)?', str(spec))
    return float(match.group()) if match else None


def decode_pcf_escaped_unicode(text):
    """Decode PCF-style escaped unicode such as \\U+6D4B into real text."""
    raw = str(text or '')
    if '\\U+' not in raw and '\\u+' not in raw:
        return raw

    def replace(match):
        hex_value = match.group(1)
        try:
            return chr(int(hex_value, 16))
        except ValueError:
            return match.group(0)

    return re.sub(r'\\[Uu]\+([0-9A-Fa-f]{4})', replace, raw)


def repair_malformed_pcf_unicode_escapes(text):
    """Repair common truncated PCF unicode escape fragments."""
    raw = str(text or '')
    replacements = (
        (r'癨U\+5F2(?=\s*头)', '°弯'),
        (r'\\[Uu]\+5F2(?=\s*头)', '弯'),
        (r'\\[Uu]\+7CF(?=\s*列)', '系'),
        (r'\\[Uu]\+56F(?=\s|$)', '图'),
        (r'(?<=\d)皠', '°'),
        (r'(?<=\d)Ёу(?=\s)', '°'),
    )
    repaired = raw
    for pattern, replacement in replacements:
        repaired = re.sub(pattern, replacement, repaired)
    return repaired


def repair_latin1_mojibake(text):
    """Repair text fragments that were persisted after latin-1 mojibake."""
    raw = str(text or '')
    mojibake_pattern = r'[\u0080-\u00ff\u0100-\u017f\u2010-\u203a]'
    if not re.search(mojibake_pattern, raw):
        return raw

    def score(value):
        cjk_count = sum(1 for char in value if '\u4e00' <= char <= '\u9fff')
        good_symbol_count = sum(value.count(symbol) for symbol in ('°', '～', '（', '）'))
        suspicious_count = sum(value.count(symbol) for symbol in ('¡', 'ã', 'Â', 'Ã', '¹', 'Ü', '¼', 'Ð', 'º', '¸', 'é', 'å', 'ç'))
        control_count = sum(1 for char in value if '\u0080' <= char <= '\u009f')
        return cjk_count * 4 + good_symbol_count * 2 - suspicious_count * 3 - control_count * 5

    def repair_chunk(match):
        chunk = match.group(0)
        if not re.search(mojibake_pattern, chunk):
            return chunk

        byte_variants = []
        try:
            byte_variants.append(chunk.encode('latin-1'))
        except UnicodeEncodeError:
            pass
        try:
            byte_variants.append(chunk.encode('cp1252'))
        except UnicodeEncodeError:
            pass

        mixed_bytes = bytearray()
        mixed_ok = True
        for char in chunk:
            try:
                mixed_bytes.extend(char.encode('cp1252'))
            except UnicodeEncodeError:
                try:
                    mixed_bytes.extend(char.encode('latin-1'))
                except UnicodeEncodeError:
                    mixed_ok = False
                    break
        if mixed_ok:
            byte_variants.append(bytes(mixed_bytes))

        best = chunk
        best_score = score(chunk)
        tried = set()
        for bytes_value in byte_variants:
            for text_encoding in ('utf-8', 'gbk', 'gb2312', 'cp936'):
                key = (bytes_value, text_encoding)
                if key in tried:
                    continue
                tried.add(key)
                try:
                    candidate = bytes_value.decode(text_encoding)
                except UnicodeDecodeError:
                    continue
                candidate_score = score(candidate)
                if candidate_score > best_score:
                    best = candidate
                    best_score = candidate_score
        return best

    chunk_pattern = r'[\x20-\x7e\u0080-\u00ff\u0100-\u017f\u2010-\u203a]*' + mojibake_pattern + r'[\x20-\x7e\u0080-\u00ff\u0100-\u017f\u2010-\u203a]*'
    return re.sub(chunk_pattern, repair_chunk, raw)


def normalize_whitespace(text):
    decoded = decode_pcf_escaped_unicode(text)
    repaired = repair_malformed_pcf_unicode_escapes(decoded)
    repaired = repair_latin1_mojibake(repaired)
    repaired = re.sub(r'\bG/T(?=\s+12459\b)', 'GB/T', repaired)
    repaired = re.sub(r'\s+', ' ', repaired).strip()
    repaired = re.sub(r'(?<=[\u4e00-\u9fff])\s+(?=[\u4e00-\u9fff])', '', repaired)
    return repaired


def join_unique_fragments(parts):
    result = []
    seen = set()
    for part in parts:
        clean = normalize_whitespace(part)
        if clean and clean not in seen:
            seen.add(clean)
            result.append(clean)
    if not result:
        return ''

    joined = result[0]
    for fragment in result[1:]:
        # Continuation DESCRIPTION lines sometimes start with punctuation,
        # e.g. "ASME B16" + ".9"; keep them tightly joined.
        if fragment[:1] in '.,;:)]}/':
            joined += fragment
        else:
            joined += ' ' + fragment
    return complete_description_text(joined.strip())


def complete_description_text(text):
    """Repair obvious truncation in normalized description text."""
    clean = normalize_whitespace(text)
    if clean.count('(') > clean.count(')') and re.search(r'\([^()]*$', clean):
        clean += ')'
    return clean


def split_pcf_sections(lines):
    """Split PCF lines into header, component area, and MATERIALS area."""
    pipeline_idx = None
    materials_idx = None

    for index, line in enumerate(lines):
        stripped = line.strip()
        if pipeline_idx is None and stripped.startswith('PIPELINE-REFERENCE'):
            pipeline_idx = index
        if stripped == 'MATERIALS':
            materials_idx = index
            break

    if pipeline_idx is None:
        pipeline_idx = 0
    if materials_idx is None:
        materials_idx = len(lines)

    component_start = materials_idx
    for index in range(pipeline_idx + 1, materials_idx):
        line = lines[index]
        stripped = line.strip()
        if not stripped:
            continue

        keyword = stripped.split(None, 1)[0]
        if line.startswith(' ') or rule_table.is_header_field_name(keyword):
            continue

        component_start = index
        break

    # Keep the entire file header, including lines that precede PIPELINE-REFERENCE
    # such as ISOGEN-FILES / UNITS-* declarations, otherwise write-back will
    # silently truncate the PCF beginning.
    header_lines = lines[:component_start]
    component_lines = lines[component_start:materials_idx]
    materials_lines = lines[materials_idx + 1:] if materials_idx < len(lines) else []
    return header_lines, component_lines, materials_lines


def load_pcf_structure(file_path):
    """Read a PCF file once and return split sections with parsed header."""
    text, _ = read_pcf_text(file_path)
    if text is None:
        return None

    lines = text.splitlines()
    header_lines, component_lines, materials_lines = split_pcf_sections(lines)
    header_meta, item_attr_map = parse_header_metadata(header_lines)
    pipeline_ref = parse_pipeline_reference(header_lines, file_path)

    return {
        'file_path': file_path,
        'lines': lines,
        'header_lines': header_lines,
        'component_lines': component_lines,
        'materials_lines': materials_lines,
        'header_meta': header_meta,
        'item_attr_map': item_attr_map,
        'pipeline_ref': pipeline_ref,
    }


def parse_pipeline_reference(header_lines, file_path):
    for line in header_lines:
        stripped = line.strip()
        if stripped.startswith('PIPELINE-REFERENCE'):
            parts = stripped.split(None, 1)
            if len(parts) > 1 and parts[1].strip():
                return parts[1].strip()
    return os.path.splitext(os.path.basename(file_path))[0]


def parse_header_metadata(header_lines):
    """Parse header metadata and item-attribute mappings."""
    metadata = {}
    item_attr_map = {}

    for line in header_lines:
        stripped = line.strip()
        if not stripped:
            continue

        parts = stripped.split(None, 1)
        key = parts[0]
        value = parts[1].strip() if len(parts) > 1 else ''

        if key.startswith('ITEM-ATTRIBUTE'):
            item_attr_map[key] = value
        metadata[key] = value

    return metadata, item_attr_map


def is_component_start(line):
    stripped = line.strip()
    return bool(stripped) and not line.startswith(' ')


def split_component_blocks(component_lines):
    """Split upper-area component section into blocks."""
    blocks = []
    current_block = []

    for line in component_lines:
        if is_component_start(line):
            if current_block:
                blocks.append(current_block)
            current_block = [line.rstrip('\n')]
        elif current_block:
            current_block.append(line.rstrip('\n'))

    if current_block:
        blocks.append(current_block)

    return blocks


def parse_point_line(line):
    stripped = line.strip()
    parts = stripped.split()
    key = parts[0]
    data = {
        'type': key,
        'raw': stripped,
        'coords': None,
        'size': None,
        'end_type': '',
    }

    if len(parts) >= 4:
        data['coords'] = tuple(parts[1:4])

    if len(parts) >= 5:
        maybe_size = safe_float_convert(parts[4])
        if maybe_size is not None:
            data['size'] = parts[4]
            if len(parts) >= 6:
                data['end_type'] = parts[5]
        else:
            data['end_type'] = parts[4]

    return data


def serialize_point(point):
    coords = point.get('coords') or ()
    coord_text = ','.join(coords) if coords else ''
    size_text = normalize_whitespace(point.get('size') or '')
    end_type_text = normalize_whitespace(point.get('end_type') or '')
    tap_text = 'TAP' if point.get('tap_connection') else ''
    parts = [normalize_whitespace(point.get('type', '')), coord_text, size_text, end_type_text, tap_text]
    return '|'.join(parts).rstrip('|')


def build_points_signature(points):
    return ' ; '.join(serialize_point(point) for point in points if point)


def build_point_types_signature(points):
    return '|'.join(normalize_whitespace(point.get('type', '')) for point in points if point)


def build_coords_signature(points):
    fragments = []
    for point in points:
        coords = point.get('coords') or ()
        if not coords:
            continue
        fragments.append(f"{normalize_whitespace(point.get('type', ''))}({','.join(coords)})")
    return ' ; '.join(fragments)


def infer_material_name(component_type, description='', item_code=''):
    description = normalize_whitespace(description).upper()
    item_code_upper = normalize_whitespace(item_code).upper()

    type_map = {
        'BEND': 'ELBOW',
        'ELBOW': 'ELBOW',
        'ELBOW-REDUCING': 'ELBOW-REDUCING',
        'PIPE': 'PIPE',
        'PIPE-FIXED': 'PIPE-FIXED',
        'FLANGE': 'FLANGE',
        'FLANGE-BLIND': 'FLANGE',
        'GASKET': 'GASKET',
        'BOLT': 'BOLT',
        'TEE': 'TEE',
        'REDUCER-ECCENTRIC': 'REDUCER-ECCENTRIC',
        'REDUCER-CONCENTRIC': 'REDUCER-CONCENTRIC',
        'OLET': 'OLET',
        'COUPLING': 'COUPLING',
        'SUPPORT': 'SUPPORT',
        'FILTER': 'FILTER',
        'CAP': 'CAP',
        'CLAMP': 'CLAMP',
    }
    if component_type in type_map:
        return type_map[component_type]

    if 'SUPPORT' in item_code_upper or 'SUPPORT' in description:
        return 'SUPPORT'
    if 'REDUCING ELBOW' in description or 'ELBOW 90 REDUING' in description or 'ELBOW REDUCING' in description:
        return 'ELBOW-REDUCING'
    if 'GASKET' in description:
        return 'GASKET'
    if 'BOLT' in description or 'NUT' in description:
        return 'BOLT'
    if 'ELB' in description:
        return 'ELBOW'
    if 'FLANGE' in description:
        return 'FLANGE'
    if 'ELBOW' in description or 'BEND' in description:
        return 'ELBOW'
    if 'VALVE' in description:
        return 'VALVE'
    if 'PIPE' in description:
        return 'PIPE'
    if 'TEE' in description:
        return 'TEE'
    if 'OLET' in description:
        return 'OLET'
    if 'COUPLING' in description:
        return 'COUPLING'
    if 'SWAGE' in description and 'ECC' in description:
        return 'REDUCER-ECCENTRIC'
    if 'SWAGE' in description:
        return 'REDUCER-CONCENTRIC'
    if 'REDUCER' in description and 'ECC' in description:
        return 'REDUCER-ECCENTRIC'
    if 'REDUCER' in description:
        return 'REDUCER-CONCENTRIC'

    return component_type


def parse_component_block(block_lines, pipeline_ref, header_meta):
    """Parse one upper-area component block."""
    component_type = block_lines[0].strip()
    item_code = ''
    description_parts = []
    specs = []
    points = []
    attributes = {}
    material = None
    material_list = None
    thickness = None
    pressure = None
    weight = None
    quantity = 1.0
    bolt_dia = None
    bolt_length = None
    bolt_quantity = None
    pending_tap_connection = False
    component_identifier = ''
    master_component_identifier = ''
    material_identifier = ''
    skey = ''
    uci = ''
    continuation = False

    for raw_line in block_lines[1:]:
        stripped = raw_line.strip()
        if not stripped:
            continue

        if stripped == 'TAP-CONNECTION':
            pending_tap_connection = True
            continue

        if stripped.startswith(('END-POINT', 'BRANCH1-POINT', 'BRANCH2-POINT', 'CENTRE-POINT', 'CO-ORDS')):
            point = parse_point_line(raw_line)
            if pending_tap_connection:
                point['tap_connection'] = True
            points.append(point)
            include_point_size = point.get('size') is not None

            # PIPE块中的 TAP-CONNECTION / CO-ORDS 代表支管坐标和支管口径，不应并入主管规格。
            if (
                component_type == 'PIPE'
                and point['type'] == 'CO-ORDS'
                and pending_tap_connection
            ):
                include_point_size = False

            if include_point_size and point['size'] not in specs:
                specs.append(point['size'])
            pending_tap_connection = False
            continue

        pending_tap_connection = False

        parts = stripped.split(None, 1)
        key = parts[0]
        value = parts[1].strip() if len(parts) > 1 else ''

        if key in {'ITEM-CODE', 'BOLT-ITEM-CODE'}:
            item_code = value
        elif key == 'COMPONENT-IDENTIFIER':
            component_identifier = value
        elif key == 'MASTER-COMPONENT-IDENTIFIER':
            master_component_identifier = value
        elif key == 'MATERIAL-IDENTIFIER':
            material_identifier = value
        elif key == 'SKEY':
            skey = value
        elif key in {'UCI', 'UNIQUE-COMPONENT-IDENTIFIER'}:
            uci = value
        elif key == 'CONTINUATION':
            continuation = True
        elif key in {'ITEM-DESCRIPTION', 'BOLT-ITEM-DESCRIPTION', 'DESCRIPTION', 'TEXT'}:
            if value:
                description_parts.append(value)
        elif key == 'WEIGHT':
            weight = safe_float_convert(value)
        elif key == 'CUT-PIECE-LENGTH':
            quantity = safe_float_convert(value) or quantity
        elif key == 'BOLT-DIA':
            bolt_dia = value
        elif key == 'BOLT-LENGTH':
            bolt_length = value
        elif key == 'BOLT-QUANTITY':
            bolt_quantity = safe_float_convert(value)
        elif key == 'PIPING-SPEC' and not attributes.get('管道等级'):
            attributes['管道等级'] = value
        elif key == 'MATERIAL-LIST':
            attributes['材料清单'] = value
            material_list = value
        elif key.startswith('COMPONENT-ATTRIBUTE'):
            attributes[key] = value
            if 'BOMCOLUMN_Material_' in value:
                material = value.split('BOMCOLUMN_Material_', 1)[1].strip()
            elif 'BOMCOLUMN_SCHClass_' in value:
                sch_value = value.split('BOMCOLUMN_SCHClass_', 1)[1].strip()
                maybe_schedule = looks_like_schedule(sch_value)
                maybe_pressure = looks_like_pressure(sch_value)
                if maybe_schedule is not None and maybe_schedule not in {'', 'X'}:
                    thickness = maybe_schedule
                elif maybe_pressure and not pressure:
                    pressure = maybe_pressure
            else:
                maybe_schedule = looks_like_schedule(value)
                maybe_pressure = looks_like_pressure(value)
                if maybe_schedule is not None and not thickness and maybe_schedule not in {'', 'X'}:
                    thickness = maybe_schedule
                elif maybe_pressure and not pressure:
                    pressure = maybe_pressure
                upper_value = value.upper()
                if pressure is None:
                    pressure_match = re.search(r'(?:CL|CLASS|LB|#)\s*([0-9]+)', upper_value)
                    if pressure_match:
                        pressure = pressure_match.group(1)
        else:
            attributes[key] = value

    description = join_unique_fragments(description_parts)

    if pressure is None:
        pressure_match = re.search(r'(?:,\s*|^)(150|300|600|900|1500|2500)(?:#|\s*LB|\s*LBS|\s*CLASS|\s*$)', description.upper())
        if pressure_match:
            pressure = pressure_match.group(1)

    if component_type == 'BOLT':
        if bolt_dia and bolt_length:
            specs = [f"{normalize_whitespace(bolt_dia)}x{normalize_whitespace(bolt_length)}"]
        if bolt_quantity is not None:
            quantity = bolt_quantity

    unit = 'PCS'
    if component_type == 'PIPE':
        end_points = [point for point in points if point['type'] == 'END-POINT' and point.get('coords')]
        if len(end_points) >= 2:
            try:
                x1, y1, z1 = map(float, end_points[0]['coords'])
                x2, y2, z2 = map(float, end_points[1]['coords'])
                quantity = round(math.dist((x1, y1, z1), (x2, y2, z2)) / 1000, 3)
                unit = 'M'
            except (TypeError, ValueError):
                if quantity:
                    quantity = round(quantity / 1000, 3)
                    unit = 'M'

    primary_points = [point for point in points if point['type'] in {'END-POINT', 'BRANCH1-POINT'}]
    start_coords = primary_points[0]['coords'] if len(primary_points) >= 1 else None
    end_coords = primary_points[1]['coords'] if len(primary_points) >= 2 else None
    points_signature = build_points_signature(points)
    point_types_signature = build_point_types_signature(points)
    coords_signature = build_coords_signature(points)

    return {
        '管线号': pipeline_ref,
        '源组件类型': component_type,
        'SKEY': skey,
        'UCI': uci,
        '组件标识': component_identifier,
        '主组件标识': master_component_identifier,
        '材料标识': material_identifier,
        '是否CONTINUATION': 'Y' if continuation else '',
        '材料名称': component_type,
        '规格列表': specs[:],
        '规格': '',
        '材料代码': item_code,
        '材料描述': description,
        '数量': round(quantity, 3) if quantity is not None else 1.0,
        '单位': unit,
        '重量': weight,
        '壁厚等级': thickness if thickness is not None else '',
        '压力等级': pressure if pressure is not None else '',
        '材质': material if material is not None else '',
        '材料清单': material_list if material_list is not None else '',
        '管道等级': attributes.get('管道等级') or header_meta.get('PIPING-SPEC', ''),
        '点位数量': len(points),
        '点位类型签名': point_types_signature,
        '坐标签名': coords_signature,
        '点位签名': points_signature,
        '点位': points,
    }


def infer_secondary_spec_from_other_components(component, all_components):
    """Infer the second spec for olet/coupling-like parts by shared coordinates."""
    if component['源组件类型'] not in {'OLET', 'COUPLING'}:
        return
    if len(component['规格列表']) >= 2:
        return

    own_specs = set(component['规格列表'])
    target_coords = {
        tuple(point['coords'])
        for point in component['点位']
        if point.get('coords')
    }

    if not target_coords:
        return

    for other in all_components:
        if other is component:
            continue
        for point in other['点位']:
            coords = point.get('coords')
            size = point.get('size')
            if not coords or not size:
                continue
            if tuple(coords) in target_coords and size not in own_specs:
                component['规格列表'].append(size)
                own_specs.add(size)
                if len(component['规格列表']) >= 2:
                    return


def infer_secondary_spec_from_raw_lines(component, all_lines):
    """Fallback: scan raw PCF lines so WELD/shared geometry lines can provide the missing size."""
    if component['源组件类型'] not in {'OLET', 'COUPLING'}:
        return
    if len(component['规格列表']) >= 2:
        return

    own_specs = set(component['规格列表'])
    target_coords = {
        tuple(point['coords'])
        for point in component['点位']
        if point.get('coords')
    }

    if not target_coords:
        return

    for raw_line in all_lines:
        stripped = raw_line.strip()
        if not stripped.startswith(('END-POINT', 'BRANCH1-POINT', 'CENTRE-POINT')):
            continue

        point = parse_point_line(raw_line)
        coords = point.get('coords')
        size = point.get('size')
        if not coords or not size:
            continue

        if tuple(coords) in target_coords and size not in own_specs:
            component['规格列表'].append(size)
            own_specs.add(size)
            if len(component['规格列表']) >= 2:
                return


def finalize_component_specs(components, all_lines=None):
    """Normalize spec list into a printable spec string."""
    for component in components:
        infer_secondary_spec_from_other_components(component, components)
        if all_lines is not None:
            infer_secondary_spec_from_raw_lines(component, all_lines)

        if component['源组件类型'] == 'BOLT':
            component['规格'] = component['规格列表'][0] if component['规格列表'] else ''
        else:
            numeric_specs = []
            original_map = {}
            for spec in component['规格列表']:
                number = extract_number(spec)
                if number is None:
                    continue
                original_map[number] = str(spec).strip()
                numeric_specs.append(number)

            numeric_specs = sorted(set(numeric_specs), reverse=True)
            if len(numeric_specs) == 1:
                number = numeric_specs[0]
                component['规格'] = str(int(number)) if float(number).is_integer() else str(number)
            elif len(numeric_specs) >= 2:
                first = str(int(numeric_specs[0])) if float(numeric_specs[0]).is_integer() else str(numeric_specs[0])
                second = str(int(numeric_specs[1])) if float(numeric_specs[1]).is_integer() else str(numeric_specs[1])
                component['规格'] = f"{first}x{second}"
            else:
                component['规格'] = ''


def parse_upper_materials(file_path, pcf_data=None):
    """Parse upper-area component instances from a PCF file."""
    if pcf_data is None:
        pcf_data = load_pcf_structure(file_path)
    if pcf_data is None:
        return []
    if 'upper_materials_cache' in pcf_data:
        return pcf_data['upper_materials_cache']

    component_lines = pcf_data['component_lines']
    header_meta = pcf_data['header_meta']
    pipeline_ref = pcf_data['pipeline_ref']

    components = []
    for block in split_component_blocks(component_lines):
        component_type = block[0].strip()
        if rule_table.is_non_material_block_type(component_type):
            continue
        component = parse_component_block(block, pipeline_ref, header_meta)
        components.append(component)

    finalize_component_specs(components, pcf_data['lines'])

    if components:
        print(f"从文件 {file_path} 中提取到 {len(components)} 个上端材料实例")
        counts = Counter(component['材料名称'] for component in components)
        print("上端材料统计:")
        for name, count in counts.items():
            print(f"  {name}: {count}个")
    else:
        print(f"文件 {file_path} 未提取到上端材料信息")

    pcf_data['upper_materials_cache'] = components
    return components


def looks_like_schedule(value):
    value = normalize_whitespace(value).upper()
    if not value or value in {'X', 'UNDEFINED'}:
        return value
    if any(token in value for token in ['SCH', 'S-', 'STD', 'XS', 'XXS']):
        return value
    return None


def looks_like_pressure(value):
    value = normalize_whitespace(value).upper()
    if not value:
        return None
    if re.fullmatch(r'\d+(?:\.\d+)?', value):
        return value
    match = re.search(r'(?:CL|CLASS|LB|#)\s*([0-9]+)', value)
    if match:
        return match.group(1)
    return None


def infer_material_name_from_lower(description, item_code):
    description_upper = normalize_whitespace(description).upper()
    code_upper = normalize_whitespace(item_code).upper()

    if 'SUPPORT' in code_upper or code_upper.startswith('J') or '管夹' in description or '支撑' in description:
        return 'SUPPORT'
    if 'INSTRUMENT' in code_upper or 'SIGHT GLASS' in description_upper:
        return 'INSTRUMENT'
    if 'PIPE' in description_upper:
        return 'PIPE'
    if 'ELB' in description_upper:
        return 'ELBOW'
    if 'ELBOW' in description_upper or 'BEND' in description_upper:
        return 'ELBOW'
    if 'FLANGE' in description_upper:
        return 'FLANGE'
    if 'GASKET' in description_upper:
        return 'GASKET'
    if 'SEAL RING' in description_upper:
        return 'GASKET'
    if 'BOLT' in description_upper or 'NUT' in description_upper or 'STUD' in description_upper:
        return 'BOLT'
    if 'TEE' in description_upper:
        return 'TEE'
    if 'OLET' in description_upper:
        return 'OLET'
    if 'COUPLING' in description_upper:
        return 'COUPLING'
    if 'FILTER' in description_upper or 'STRAINER' in description_upper:
        return 'FILTER'
    if 'VALVE' in description_upper:
        return 'VALVE'
    if 'CAP' in description_upper:
        return 'CAP'
    if 'SWAGE' in description_upper and 'ECC' in description_upper:
        return 'REDUCER-ECCENTRIC'
    if 'SWAGE' in description_upper:
        return 'REDUCER-CONCENTRIC'
    if 'REDUCER' in description_upper and 'ECC' in description_upper:
        return 'REDUCER-ECCENTRIC'
    if 'REDUCER' in description_upper:
        return 'REDUCER-CONCENTRIC'
    if code_upper.startswith('BUTTWELD'):
        return 'WELD-MATERIAL'
    if code_upper.startswith('TAPWELD'):
        return 'WELD-MATERIAL'
    return ''


def infer_material_name_from_code_pattern(item_code):
    code_upper = normalize_whitespace(item_code).upper()

    if not code_upper or code_upper in {'..'}:
        return ''
    if code_upper.startswith('BUTTWELD') or code_upper.startswith('TAPWELD'):
        return 'WELD-MATERIAL'
    if re.match(r'^(45|90)?EL', code_upper):
        return 'ELBOW'
    if code_upper.startswith('RC'):
        return 'REDUCER-CONCENTRIC'
    if code_upper.startswith(('RE', 'ECC')):
        return 'REDUCER-ECCENTRIC'
    if re.match(r'^P\d', code_upper):
        return 'PIPE'
    if re.match(r'^F\d', code_upper):
        return 'FLANGE'
    if re.match(r'^O\d', code_upper):
        return 'OLET'
    if re.match(r'^T\d+S', code_upper):
        return 'TEE'
    if code_upper.startswith('G'):
        return 'GASKET'
    if code_upper.startswith('B'):
        return 'BOLT'
    return ''


def align_material_groups_with_codes(pending_codes, pending_groups):
    """Reorder parsed MATERIALS groups so code-pattern expectations match descriptions better."""
    if len(pending_codes) <= 1 or len(pending_codes) != len(pending_groups) or len(pending_codes) > 6:
        return pending_groups

    expected_names = [infer_material_name_from_code_pattern(code) for code in pending_codes]
    group_names = [
        infer_material_name_from_lower(join_unique_fragments(group.get('DESCRIPTION', [])), '')
        for group in pending_groups
    ]

    if not any(expected_names) or not any(group_names):
        return pending_groups

    def score_pair(expected_name, group_name):
        if expected_name and group_name:
            return 3 if expected_name == group_name else -2
        if expected_name and not group_name:
            return 1
        return 0

    best_order = list(range(len(pending_groups)))
    best_score = sum(score_pair(expected_names[index], group_names[index]) for index in range(len(pending_groups)))

    for order in permutations(range(len(pending_groups))):
        score = sum(score_pair(expected_names[index], group_names[order[index]]) for index in range(len(pending_groups)))
        if score > best_score:
            best_score = score
            best_order = list(order)

    return [pending_groups[index] for index in best_order]


def enrich_lower_material_names_from_upper(results, upper_materials):
    """Backfill lower material names from upper parsed instances when the mapping is unique."""
    code_to_names = defaultdict(set)
    desc_to_names = defaultdict(set)

    for item in upper_materials:
        material_name = normalize_whitespace(item.get('材料名称', ''))
        code = normalize_whitespace(item.get('材料代码', ''))
        description = normalize_whitespace(item.get('材料描述', ''))

        if material_name:
            if code:
                code_to_names[code].add(material_name)
            if description:
                desc_to_names[description].add(material_name)

    for entry in results:
        if entry.get('材料名称'):
            continue

        code = normalize_whitespace(entry.get('材料代码', ''))
        description = normalize_whitespace(entry.get('材料描述', ''))

        code_names = code_to_names.get(code, set())
        desc_names = desc_to_names.get(description, set())

        if len(code_names) == 1:
            entry['材料名称'] = next(iter(code_names))
            continue

        if len(desc_names) == 1 and not code_names:
            entry['材料名称'] = next(iter(desc_names))

    return results


def map_material_user_fields(entry):
    """Map variant MATERIAL-USER fields into normalized lower material fields."""
    thickness = normalize_whitespace(entry.get('壁厚等级', ''))
    pressure = normalize_whitespace(entry.get('压力等级', ''))
    material = normalize_whitespace(entry.get('材质', ''))
    material_list = normalize_whitespace(entry.get('材料清单', ''))

    for key, value in entry.get('原始字段', {}).items():
        clean_value = normalize_whitespace(value)
        if not clean_value:
            continue

        if key in {'Schedule', 'MATERIAL-USER5'} and not thickness:
            maybe = looks_like_schedule(clean_value)
            if maybe is not None:
                thickness = maybe
        elif key == 'MATERIAL-LIST' and not material_list:
            material_list = clean_value
        elif key == 'PressureClass' and not pressure:
            pressure = clean_value
        elif key == 'Material' and not material:
            material = clean_value
        elif key.startswith('MATERIAL-USER'):
            if not thickness:
                maybe = looks_like_schedule(clean_value)
                if maybe is not None:
                    thickness = maybe
            if not pressure:
                maybe_pressure = looks_like_pressure(clean_value)
                if maybe_pressure and maybe_pressure not in {'0', '0.0'}:
                    pressure = maybe_pressure

    entry['壁厚等级'] = '' if thickness == 'X' else thickness
    entry['压力等级'] = pressure
    entry['材质'] = material
    entry['材料清单'] = material_list
    return entry


def flush_material_entries(pending_codes, pending_groups, pipeline_ref, results):
    if not pending_codes and not pending_groups:
        return

    if not pending_groups:
        pending_groups = [{} for _ in pending_codes]

    if len(pending_groups) == 1 and len(pending_codes) > 1:
        pending_groups = pending_groups * len(pending_codes)

    pending_groups = align_material_groups_with_codes(pending_codes, pending_groups)

    max_len = max(len(pending_codes), len(pending_groups))

    for index in range(max_len):
        outer_code = pending_codes[index] if index < len(pending_codes) else ''
        group = pending_groups[index] if index < len(pending_groups) else {}
        nested_code = group.get('ITEM-CODE', outer_code)
        if isinstance(nested_code, list):
            nested_code = join_unique_fragments(nested_code)
        actual_code = normalize_whitespace(nested_code or outer_code)
        description = join_unique_fragments(group.get('DESCRIPTION', []))

        entry = {
            '管线号': pipeline_ref,
            '材料标识': normalize_whitespace(outer_code),
            '材料代码': actual_code,
            '材料描述': description,
            '壁厚等级': '',
            '压力等级': '',
            '材质': '',
            '材料清单': '',
            '材料名称': infer_material_name_from_lower(description, actual_code),
            '原始外层代码': outer_code,
            '原始字段': {},
        }

        for key, value in group.items():
            if key in {'DESCRIPTION', 'ITEM-CODE'}:
                continue
            entry['原始字段'][key] = join_unique_fragments(value) if isinstance(value, list) else normalize_whitespace(value)

        entry = map_material_user_fields(entry)
        results.append(entry)


def parse_materials_section(file_path, pcf_data=None):
    """Parse the MATERIALS section across different project dialects."""
    if pcf_data is None:
        pcf_data = load_pcf_structure(file_path)
    if pcf_data is None:
        return []

    materials_lines = pcf_data['materials_lines']
    pipeline_ref = pcf_data['pipeline_ref']

    results = []
    pending_codes = []
    pending_groups = []
    current_group = None

    def ensure_group():
        nonlocal current_group
        if current_group is None:
            current_group = defaultdict(list)
            pending_groups.append(current_group)
        return current_group

    for raw_line in materials_lines:
        line = raw_line.rstrip()
        stripped = line.strip()
        if not stripped:
            continue

        outer_key = ''
        if not line.startswith(' '):
            parts = stripped.split(None, 1)
            outer_key = parts[0] if parts else ''

        if outer_key in {'ITEM-CODE', 'MATERIAL-IDENTIFIER'}:
            code = normalize_whitespace(stripped.split(outer_key, 1)[1])
            if pending_codes and pending_groups:
                flush_material_entries(pending_codes, pending_groups, pipeline_ref, results)
                pending_codes = []
                pending_groups = []
                current_group = None
            pending_codes.append(code)
            continue

        if line.startswith(' ') and pending_codes:
            parts = stripped.split(None, 1)
            key = parts[0]
            value = parts[1].strip() if len(parts) > 1 else ''

            if key == 'DESCRIPTION':
                if current_group and current_group.get('DESCRIPTION') and len(current_group.keys()) > 1:
                    current_group = None
                ensure_group()['DESCRIPTION'].append(value)
            elif key == 'ITEM-CODE':
                if current_group and current_group.get('ITEM-CODE'):
                    current_group = None
                ensure_group()['ITEM-CODE'].append(value)
            else:
                group = ensure_group()
                if key in group and key != 'DESCRIPTION':
                    current_group = None
                    group = ensure_group()
                group[key].append(value)

    flush_material_entries(pending_codes, pending_groups, pipeline_ref, results)

    cleaned_results = []
    for entry in results:
        entry['材料名称'] = entry['材料名称'] or infer_material_name_from_lower(entry['材料描述'], entry['材料代码'])
        entry.pop('原始字段', None)
        entry.pop('原始外层代码', None)
        if any(entry.get(field, '') for field in ['材料代码', '材料描述', '壁厚等级', '压力等级', '材质']):
            cleaned_results.append(entry)

    upper_materials = parse_upper_materials(file_path, pcf_data)
    enrich_lower_material_names_from_upper(cleaned_results, upper_materials)

    if cleaned_results:
        print(f"从文件 {file_path} 中提取到 {len(cleaned_results)} 条下端材料定义")
    else:
        print(f"文件 {file_path} 未找到有效的 MATERIALS 材料定义")

    return cleaned_results


def format_worksheet(sheet):
    """Format a worksheet in-place before workbook save."""
    if not ENABLE_EXCEL_FORMATTING:
        return

    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    header_fill = PatternFill(start_color='E6FFE6', end_color='E6FFE6', fill_type='solid')
    font = Font(name='微软雅黑', size=9)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )
    alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    for row in range(1, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=row, column=col)
            cell.font = font
            cell.border = thin_border
            cell.alignment = alignment
            if row == 1:
                cell.fill = header_fill

    for col in range(1, sheet.max_column + 1):
        max_length = 0
        column = get_column_letter(col)
        for row in range(1, sheet.max_row + 1):
            value = sheet.cell(row=row, column=col).value
            text = '' if value is None else str(value)
            max_length = max(max_length, len(text))
        sheet.column_dimensions[column].width = max_length + 4


def discover_pcf_files(root_dir='.'):
    """Recursively discover all PCF files under the target directory."""
    pcf_files = []

    for current_root, dir_names, file_names in os.walk(root_dir):
        dir_names.sort()
        for file_name in sorted(file_names):
            if file_name.lower().endswith('.pcf'):
                pcf_files.append(os.path.join(current_root, file_name))

    return pcf_files


def get_unit_name(file_path):
    """Use the containing folder name as the unit identifier."""
    parent_dir = os.path.basename(os.path.dirname(os.path.abspath(file_path)))
    return parent_dir or os.path.basename(os.path.abspath('.'))


def attach_unit_name(records, file_path):
    """Add unit identifier to parsed records."""
    unit_name = get_unit_name(file_path)
    for record in records:
        record['单元号'] = unit_name
    return records


def enrich_upper_materials_from_lower(upper_materials, lower_materials):
    """Attach lower MATERIALS code/description to upper component rows."""
    by_identifier = {}
    by_code = {}

    for item in lower_materials:
        identifier = normalize_whitespace(item.get('材料标识', ''))
        code = normalize_whitespace(item.get('材料代码', ''))
        description = normalize_whitespace(item.get('材料描述', ''))

        if identifier and identifier not in by_identifier:
            by_identifier[identifier] = item
        if code and code not in by_code:
            by_code[code] = item
        item.setdefault('下端材料描述', description)

    for item in upper_materials:
        identifier = normalize_whitespace(item.get('材料标识', ''))
        code = normalize_whitespace(item.get('材料代码', ''))
        lower_item = None

        if identifier:
            lower_item = by_identifier.get(identifier)
        if lower_item is None and code:
            lower_item = by_code.get(code)

        if lower_item is None:
            item.setdefault('下端材料描述', '')
            continue

        lower_code = normalize_whitespace(lower_item.get('材料代码', ''))
        lower_description = normalize_whitespace(lower_item.get('材料描述', ''))

        if lower_code and (not code or code == identifier):
            item['材料代码'] = lower_code
        item['下端材料描述'] = lower_description
        if lower_description and not normalize_whitespace(item.get('材料描述', '')):
            item['材料描述'] = lower_description

    return upper_materials


def main():
    start_folder = os.environ.get('PIPECLOUD_PARSER_INPUT_DIR') or str(DEFAULT_FILE_DIR)
    output_folder = os.environ.get('PIPECLOUD_PARSER_OUTPUT_DIR') or start_folder
    os.makedirs(output_folder, exist_ok=True)
    pcf_files = discover_pcf_files(start_folder)
    print(f"Found {len(pcf_files)} PCF files.")

    if not pcf_files:
        print('当前目录及其子目录下没有找到PCF文件')
        return

    all_lower_materials = []
    all_upper_materials = []

    for file_name in pcf_files:
        print(f"\n正在处理文件: {file_name}")
        pcf_data = load_pcf_structure(file_name)
        if pcf_data is None:
            continue

        lower_materials = parse_materials_section(file_name, pcf_data)
        upper_materials = parse_upper_materials(file_name, pcf_data)
        enrich_upper_materials_from_lower(upper_materials, lower_materials)

        lower_materials = attach_unit_name(lower_materials, file_name)
        upper_materials = attach_unit_name(upper_materials, file_name)
        all_lower_materials.extend(lower_materials)
        all_upper_materials.extend(upper_materials)

    excel_file = os.path.join(output_folder, 'PCF材料段输出信息.xlsx')

    with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
        lower_columns = ['单元号', '管线号', '材料代码', '材料描述', '材料名称', '壁厚等级', '压力等级', '材质', '材料清单']
        upper_columns = [
            '单元号', '管线号', 'SKEY', 'UCI', '组件标识', '主组件标识', '材料标识', '是否CONTINUATION',
            '点位数量', '点位类型签名', '坐标签名', '点位签名',
            '规格', '材料代码', '材料描述', '下端材料描述', '数量', '单位', '材料名称', '重量',
            '壁厚等级', '压力等级', '材质', '材料清单', '管道等级', '修改后材料代码'
        ]

        if all_lower_materials:
            lower_df = pd.DataFrame(all_lower_materials)
            for column in lower_columns:
                if column not in lower_df.columns:
                    lower_df[column] = ''
            lower_df = lower_df[lower_columns]
            lower_df.to_excel(writer, sheet_name='下端数据读取', index=False)
        else:
            pd.DataFrame(columns=lower_columns).to_excel(writer, sheet_name='下端数据读取', index=False)

        if all_upper_materials:
            upper_df = pd.DataFrame(all_upper_materials)
            for column in upper_columns:
                if column not in upper_df.columns:
                    upper_df[column] = ''
            upper_df = upper_df[upper_columns]
            upper_df.to_excel(writer, sheet_name='上端数据读取', index=False)
        else:
            pd.DataFrame(columns=upper_columns).to_excel(writer, sheet_name='上端数据读取', index=False)
        if ENABLE_EXCEL_FORMATTING:
            format_worksheet(writer.sheets['下端数据读取'])
            format_worksheet(writer.sheets['上端数据读取'])

    if ENABLE_EXCEL_FORMATTING:
        print(f'材料信息已保存到 {excel_file}，已包含格式处理')
    else:
        print(f'材料信息已快速保存到 {excel_file}，未进行格式处理')


if __name__ == '__main__':
    main()

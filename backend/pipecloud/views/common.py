import json
import locale
import os
import re
import shutil
import subprocess
import sys
import tempfile
import uuid
from decimal import Decimal, InvalidOperation
from datetime import datetime, timedelta
from io import BytesIO
from pathlib import Path
from urllib.parse import quote

import pandas as pd
from django.core.exceptions import ObjectDoesNotExist
from django.db.models import Count
from django.http import HttpResponse, HttpResponseBadRequest, JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_GET, require_POST
from openpyxl import load_workbook
from spool_analysis.project_spool_info import (
    read_project_spool_info_from_database,
)

from ..models import (
    ArrivalMaterialRow,
    ArrivalOrderRow,
    DataSourceFile,
    FittingMaterialRow,
    InitializationWeldRow,
    MasterScheduleRow,
    PipeMaterialRow,
    PlanRecord,
    Project,
    ProjectSchedulePolicy,
    WeldingPlanRow,
    WeldLibraryRow,
    WeldPreScheduleRow,
)
from pipecloud.services.project_tables import ensure_project_tables, using_project_tables
from pipecloud.services.project_constraints import project_process_sequence
from pipecloud.services.db_storage import PRE_SCHEDULE_MODELS, replace_source_rows, table_payload


BACKEND_DIR = Path(__file__).resolve().parent.parent.parent
PREFAB_ROOT = BACKEND_DIR / 'prefab_schedule'
SPOOL_ANALYSIS_ROOT = BACKEND_DIR / 'spool_analysis'
FILE_ROOT = BACKEND_DIR / 'file'
PROJECTS_ROOT = FILE_ROOT / 'projects'
FILE_PARSER_ROOT = FILE_ROOT / 'parser'
FILE_BACKUP_ROOT = FILE_ROOT / 'backups'
DATA_ROOT = PROJECTS_ROOT
STAGED_PLAN_ROOT = FILE_ROOT / 'staged_plans'
LIBRARY_ROOT = DATA_ROOT / '库管理'
ARRIVAL_ROOT = DATA_ROOT / '入库单'
CUTTING_PIPE_LIBRARY = LIBRARY_ROOT / '防腐管子材料库.xlsx'
CUTTING_PRE_SCHEDULE_OUTPUT = DATA_ROOT / '中间结果' / '焊口预排产匹配结果.xlsx'
ANTI_CORROSION_PLAN_ROOT = DATA_ROOT / '防腐委托单'
CUTTING_PLAN_ROOT = DATA_ROOT / '下料排产单'
WELDING_PLAN_ROOT = DATA_ROOT / '焊接排产单'

PROJECT_COLUMNS = [
    {'field': 'project_name', 'title': '项目名称'},
    {'field': 'pipe_segment', 'title': '可预制管段'},
    {'field': 'prefab_weld_count', 'title': '可预制焊口数'},
    {'field': 'completion_rate', 'title': '完成率'},
    {'field': 'start_date', 'title': '开始日期'},
]
PROJECT_FIELD_NAMES = {column['field'] for column in PROJECT_COLUMNS}
PROJECT_TITLE_TO_FIELD = {column['title']: column['field'] for column in PROJECT_COLUMNS}
PROJECT_TITLE_TO_FIELD['管段'] = 'pipe_segment'
PROJECT_DATA_ROOT = PROJECTS_ROOT
INIT_WELD_FILE_PATTERN = '焊口初始化数据*.xlsx'
PARSER_OUTPUT_FILES = {
    'idf': 'IDF拓扑材料表.xlsx',
    'pcf': 'PCF材料段输出信息.xlsx',
}
PARSER_SCRIPTS = {
    'idf': SPOOL_ANALYSIS_ROOT / 'parse_idf_viewer_to_excel.py',
    'pcf': SPOOL_ANALYSIS_ROOT / 'rename_pcf_files_v5.py',
}

CUTTING_COLUMNS = {
    'material_code': '材料代码',
    'description': '名称及规格',
    'pipe_no': '管子序号',
    'stock_qty': '库存数量（米）',
    'original_length': '原始米数',
    'remaining_length': '剩余米数',
    'cut_lengths': '已切割长度列表',
    'loss_lengths': '切割损耗列表',
    'consumed_lengths': '实际占用长度列表',
    'heat_no': '炉批号',
}

WELD_LIBRARY_FILE_NAME = '预制焊口库.xlsx'
WELDING_PRIMARY_PLAN_FILE_NAME = '管段焊口表.xlsx'
WELD_COLUMNS = {
    'unit': '单元号',
    'pipeline': '管线号',
    'segment_no': '管段号',
    'diameter': '寸径',
    'weld_no_start': '初始焊口号',
    'weld_no_final': '最终焊口号',
    'library_seq': '库序号',
    'completed_flag': '材料焊接状态',
    'material_anti_corrosion_status': '材料防腐状态',
}
WELD_COLUMN_ALIASES = {
    WELD_COLUMNS['segment_no']: ['预制组件', '管段号', '预制管段', '预制段'],
    WELD_COLUMNS['completed_flag']: ['材料焊接状态', '是否完成', '完成状态', '状态', '完工状态', '焊接状态'],
    WELD_COLUMNS['material_anti_corrosion_status']: ['材料防腐状态', '防腐状态', '是否防腐完成', '防腐完成状态'],
    WELD_COLUMNS['weld_no_start']: ['初始焊口号', '起始焊口号', '焊口起始号', '开始焊口号'],
    WELD_COLUMNS['weld_no_final']: ['最终焊口号', '结束焊口号', '焊口结束号', '完成焊口号'],
    WELD_COLUMNS['pipeline']: ['管线号(必填)', '管线号'],
    WELD_COLUMNS['unit']: ['单元号(必填)', '单元号'],
    WELD_COLUMNS['diameter']: ['英制', '英制尺寸', '寸径'],
    WELD_COLUMNS['library_seq']: ['库序号', '序号', '行号'],
}

PLAN_SOURCES = {
    'anti-corrosion': {
        'key': 'anti-corrosion',
        'name': '防腐',
        'root': ANTI_CORROSION_PLAN_ROOT,
    },
    'cutting': {
        'key': 'cutting',
        'name': '下料',
        'root': CUTTING_PLAN_ROOT,
    },
    'welding': {
        'key': 'welding',
        'name': '焊接',
        'root': WELDING_PLAN_ROOT,
    },
}


MODULES = [
    {
        'key': 'initialization',
        'name': '初始化预制',
        'description': '读取焊口初始化数据，完成自动焊划分，并生成预制焊口库。',
        'actions': ['prefab-weld-library'],
        'files': [
            '焊口初始化数据*.xlsx',
            '中间结果/可预制焊口初步过滤结果.xlsx',
            '中间结果/自动焊口初步过滤结果.xlsx',
            '中间结果/链路划分结果.xlsx',
            '中间结果/自动焊口数据.xlsx',
            '库管理/预制焊口库.xlsx',
        ],
    },
    {
        'key': 'arrival',
        'name': '到货管理',
        'description': '汇总入库单，维护管子材料库和管件法兰材料库。',
        'actions': [],
        'files': [
            '入库单',
            '库管理/管子材料库.xlsx',
            '库管理/管件法兰材料库.xlsx',
        ],
    },
    {
        'key': 'materialLocking',
        'name': '材料匹配与锁定',
        'description': '按项目约束匹配并锁定焊口材料，更新材料到货状态并展示管子切割占用。',
        'actions': ['material-locking'],
        'files': [
            '库管理/管子材料库.xlsx',
            '库管理/管件法兰材料库.xlsx',
            '中间结果/材料匹配锁定结果.xlsx',
        ],
    },
    {
        'key': 'antiCorrosion',
        'name': '防腐管理及排产',
        'description': '根据材料锁定结果生成需防腐焊口的防腐预排产和防腐委托。',
        'actions': ['anti-corrosion-pre-schedule', 'anti-corrosion-schedule'],
        'files': [
            '库管理/防腐管子材料库.xlsx',
            '库管理/防腐管件法兰材料库.xlsx',
            '中间结果/防腐预排产匹配结果.xlsx',
        ],
    },
    {
        'key': 'cutting',
        'name': '下料管理及排产',
        'description': '提取材料已到货、已防腐且尚未下料的焊口，生成下料预排产和下料排产单。',
        'actions': ['weld-pre-schedule', 'cutting-schedule'],
        'files': [
            '中间结果/焊口预排产匹配结果.xlsx',
        ],
    },
    {
        'key': 'welding',
        'name': '焊接管理及排产',
        'description': '在初始化预制和下料预排产确认完成后，生成焊接预排产和焊接排产单。',
        'actions': ['welding-pre-schedule', 'auto-weld-schedule'],
        'files': [
            '中间结果/焊接预排产结果.xlsx',
            '焊接排产单',
        ],
    },
    {
        'key': 'schedule',
        'name': '总排产计划',
        'description': '按预排产结果滚动生成未来下料计划、焊接计划和总排产计划。',
        'actions': ['future-schedule'],
        'files': [
            '中间结果/焊口预排产匹配结果.xlsx',
            '下料排产单',
            '焊接排产单',
            '焊接排产单/总排产计划.xlsx',
        ],
    },
]


def _modules_for_project(project):
    modules = [dict(module) for module in MODULES]
    if project is not None and project_process_sequence(project) == 'welding_before_coating':
        order = {module['key']: index for index, module in enumerate(modules)}
        order['antiCorrosion'] = order.get('welding', order.get('antiCorrosion', 0)) + 0.5
        modules.sort(key=lambda module: order.get(module['key'], 999))
    return modules


ACTIONS = {
    'prefab-weld-library': {
        'name': '生成预制焊口库',
        'script': PREFAB_ROOT / 'initialization' / 'build_prefab_weld_library.py',
        'module': '初始化预制',
    },
    'arrival-library': {
        'name': '生成材料库',
        'script': PREFAB_ROOT / 'arrival' / 'material_library_maintenance.py',
        'module': '到货管理',
    },
    'update-weld-arrival-status': {
        'name': '更新预制焊口库材料到货状态',
        'script': PREFAB_ROOT / 'arrival' / 'material_library_maintenance.py',
        'module': '到货管理',
    },
    'material-locking': {
        'name': '材料匹配与锁定',
        'script': PREFAB_ROOT / 'cutting' / 'weld_pre_schedule_matcher.py',
        'module': '材料匹配与锁定',
    },
    'anti-corrosion-schedule': {
        'name': '生成防腐委托',
        'script': PREFAB_ROOT / 'anti_corrosion' / 'main.py',
        'module': '防腐管理及排产',
    },
    'anti-corrosion-pre-schedule': {
        'name': '生成防腐预排产',
        'script': PREFAB_ROOT / 'anti_corrosion' / 'pre_schedule_matcher.py',
        'module': '防腐管理及排产',
    },
    'auto-weld-schedule': {
        'name': '生成焊接排产单',
        'script': PREFAB_ROOT / 'welding' / 'auto_weld_schedule' / 'main.py',
        'module': '焊接管理及排产',
    },
    'welding-pre-schedule': {
        'name': '生成焊接预排产',
        'script': PREFAB_ROOT / 'welding' / 'main.py',
        'module': '焊接管理及排产',
    },
    'weld-pre-schedule': {
        'name': '生成下料预排产',
        'script': PREFAB_ROOT / 'cutting' / 'weld_pre_schedule_matcher.py',
        'module': '下料管理及排产',
    },
    'cutting-schedule': {
        'name': '生成下料排产单',
        'script': PREFAB_ROOT / 'cutting' / 'main.py',
        'module': '下料管理及排产',
    },
    'future-schedule': {
        'name': '生成所有计划',
        'script': PREFAB_ROOT / 'schedule.py',
        'module': '总排产计划',
    },
}


LIBRARY_FILES = {
    'weld-library': {
        'name': '预制焊口库',
        'path': LIBRARY_ROOT / '预制焊口库.xlsx',
    },
    'pipe-library': {
        'name': '管子材料库',
        'path': LIBRARY_ROOT / '管子材料库.xlsx',
    },
    'fitting-library': {
        'name': '管件法兰材料库',
        'path': LIBRARY_ROOT / '管件法兰材料库.xlsx',
    },
    'anti-pipe-library': {
        'name': '防腐管子材料库',
        'path': LIBRARY_ROOT / '防腐管子材料库.xlsx',
    },
    'anti-fitting-library': {
        'name': '防腐管件法兰材料库',
        'path': LIBRARY_ROOT / '防腐管件法兰材料库.xlsx',
    },
}


def _relative_path(path):
    try:
        return str(path.relative_to(BACKEND_DIR)).replace('\\', '/')
    except ValueError:
        pass
    try:
        return str(path.relative_to(PREFAB_ROOT)).replace('\\', '/')
    except ValueError:
        return str(path).replace('\\', '/')


def _request_project_context(request, required=False):
    project_id = request.GET.get('project_id') or request.POST.get('project_id')
    if not project_id:
        return None, DATA_ROOT, '请先选择项目' if required else ''
    try:
        project = Project.objects.get(pk=int(project_id))
    except (TypeError, ValueError, Project.DoesNotExist):
        return None, DATA_ROOT, '项目不存在'
    return project, _project_root(project), ''


def _project_bad_request(error):
    return HttpResponseBadRequest(
        json.dumps({'error': error}, ensure_ascii=False),
        content_type='application/json',
    )


def _library_files(data_root):
    library_root = data_root / '库管理'
    return {
        'weld-library': {
            'name': '预制焊口库',
            'path': library_root / '预制焊口库.xlsx',
        },
        'pipe-library': {
            'name': '管子材料库',
            'path': library_root / '管子材料库.xlsx',
        },
        'fitting-library': {
            'name': '管件法兰材料库',
            'path': library_root / '管件法兰材料库.xlsx',
        },
        'anti-pipe-library': {
            'name': '防腐管子材料库',
            'path': library_root / '防腐管子材料库.xlsx',
        },
        'anti-fitting-library': {
            'name': '防腐管件法兰材料库',
            'path': library_root / '防腐管件法兰材料库.xlsx',
        },
    }


def _plan_sources(data_root):
    return {
        'anti-corrosion': {
            'key': 'anti-corrosion',
            'name': '防腐',
            'root': data_root / '防腐委托单',
        },
        'cutting': {
            'key': 'cutting',
            'name': '下料',
            'root': data_root / '下料排产单',
        },
        'welding': {
            'key': 'welding',
            'name': '焊接',
            'root': data_root / '焊接排产单',
        },
    }


def _database_status_payload(relative_name, exists=False, count=0, updated_at=None, path=None, status_type='数据库'):
    return {
        'name': relative_name,
        'path': path or relative_name,
        'exists': bool(exists),
        'type': status_type,
        'count': int(count or 0),
        'size': None,
        'updatedAt': updated_at or None,
    }


def _latest_source_from_queryset(queryset):
    return queryset.order_by('-file_updated_at', '-id').first()


def _database_source_file_info(
    project,
    relative_name,
    source_type=None,
    source_key=None,
    source_key_prefix=None,
    display_name=None,
    row_model=None,
    row_filters=None,
    source_count_only=False,
):
    if project is None:
        return _database_status_payload(relative_name)

    source_queryset = DataSourceFile.objects.filter(project=project)
    if source_type:
        source_queryset = source_queryset.filter(source_type=source_type)
    if source_key:
        source_queryset = source_queryset.filter(source_key=source_key)
    if source_key_prefix:
        source_queryset = source_queryset.filter(source_key__startswith=source_key_prefix)
    if display_name:
        source_queryset = source_queryset.filter(display_name=display_name)

    source_count = source_queryset.count()
    latest_source = _latest_source_from_queryset(source_queryset)
    row_count = None
    if row_model is not None and not source_count_only:
        row_queryset = row_model.objects.filter(project=project)
        if source_type:
            row_queryset = row_queryset.filter(source_file__source_type=source_type)
        if source_key:
            row_queryset = row_queryset.filter(source_file__source_key=source_key)
        if source_key_prefix:
            row_queryset = row_queryset.filter(source_file__source_key__startswith=source_key_prefix)
        if display_name:
            row_queryset = row_queryset.filter(source_file__display_name=display_name)
        if row_filters:
            row_queryset = row_queryset.filter(**row_filters)
        row_count = row_queryset.count()

    count = source_count if source_count_only or row_count is None else row_count
    return _database_status_payload(
        relative_name,
        exists=source_count > 0 or (row_count or 0) > 0,
        count=count,
        updated_at=latest_source.file_updated_at if latest_source else None,
        path=latest_source.relative_path if latest_source else relative_name,
    )


def _database_plan_record_info(project, relative_name, plan_key, row_model=None):
    if project is None:
        return _database_status_payload(relative_name)
    records = PlanRecord.objects.filter(project=project, plan_key=plan_key)
    latest = records.order_by('-folder_updated_at', '-id').first()
    row_count = 0
    if row_model is not None:
        row_count = row_model.objects.filter(
            project=project,
            source_file__source_type='plan',
            source_file__source_key__startswith=f'{plan_key}:',
        ).count()
    count = row_count or records.count()
    return _database_status_payload(
        relative_name,
        exists=records.exists() or row_count > 0,
        count=count,
        updated_at=latest.folder_updated_at if latest else None,
        path=latest.relative_path if latest else relative_name,
        status_type='计划记录',
    )


def _database_module_file_info(project, relative_name, data_root=DATA_ROOT):
    if project is None:
        return _database_status_payload(relative_name)

    with using_project_tables(project):
        if relative_name == '焊口初始化数据*.xlsx':
            return _database_source_file_info(project, relative_name, 'initialization', 'welds', row_model=InitializationWeldRow)
        if relative_name in {'库管理/焊口库.xlsx', '库管理/预制焊口库.xlsx'}:
            return _database_source_file_info(project, relative_name, 'library', 'weld-library', row_model=WeldLibraryRow)
        if relative_name in {'中间结果/可预制焊口初步过滤结果.xlsx', '中间结果/链路划分结果.xlsx'}:
            return _database_source_file_info(project, relative_name, 'library', 'weld-library', row_model=WeldLibraryRow)
        if relative_name in {'中间结果/自动焊口初步过滤结果.xlsx', '中间结果/自动焊口数据.xlsx'}:
            return _database_source_file_info(
                project,
                relative_name,
                'library',
                'weld-library',
                row_model=WeldLibraryRow,
                row_filters={'welding_mode__contains': '自动焊'},
            )
        if relative_name == '入库单':
            return _database_source_file_info(project, relative_name, 'arrival', row_model=ArrivalOrderRow, source_count_only=True)
        if relative_name == '库管理/管子材料库.xlsx':
            return _database_source_file_info(project, relative_name, 'library', 'pipe-library', row_model=PipeMaterialRow)
        if relative_name == '库管理/管件法兰材料库.xlsx':
            return _database_source_file_info(project, relative_name, 'library', 'fitting-library', row_model=FittingMaterialRow)
        if relative_name == '库管理/防腐管子材料库.xlsx':
            return _database_source_file_info(project, relative_name, 'library', 'anti-pipe-library', row_model=PipeMaterialRow)
        if relative_name == '库管理/防腐管件法兰材料库.xlsx':
            return _database_source_file_info(project, relative_name, 'library', 'anti-fitting-library', row_model=FittingMaterialRow)
        if relative_name == '防腐委托单/防腐委托总表.xlsx':
            return _database_plan_record_info(project, relative_name, 'anti-corrosion')
        if relative_name == '中间结果/防腐预排产匹配结果.xlsx':
            return _database_source_file_info(project, relative_name, 'pre-schedule', 'anti-corrosion-pre-schedule', row_model=WeldPreScheduleRow)
        if relative_name == '中间结果/材料匹配锁定结果.xlsx':
            return _database_source_file_info(project, relative_name, 'pre-schedule', 'material-locking', row_model=WeldPreScheduleRow)
        if relative_name == '中间结果/焊接预排产结果.xlsx':
            return _database_source_file_info(project, relative_name, 'pre-schedule', 'welding-pre-schedule', row_model=WeldPreScheduleRow)
        if relative_name == '中间结果/焊口预排产匹配结果.xlsx':
            return _database_source_file_info(project, relative_name, 'pre-schedule', 'weld-pre-schedule', row_model=WeldPreScheduleRow)
        if relative_name == '下料排产单':
            return _database_plan_record_info(project, relative_name, 'cutting')
        if relative_name == '焊接排产单':
            return _database_plan_record_info(project, relative_name, 'welding', WeldingPlanRow)
        if relative_name == '焊接排产单/总排产计划.xlsx':
            return _database_plan_record_info(project, relative_name, 'welding', WeldingPlanRow)
    return _database_status_payload(relative_name)


def _clean_cell(value):
    if pd.isna(value):
        return ''
    if hasattr(value, 'isoformat'):
        return value.isoformat()
    return value


def _project_payload(project):
    weld_source = _latest_source(project, 'initialization', 'welds')
    prefab_source = _latest_source(project, 'library', 'weld-library')
    return {
        'id': project.id,
        'project_name': project.project_name,
        'pipe_segment': project.pipe_segment,
        'prefab_weld_count': project.prefab_weld_count,
        'completion_rate': '' if project.completion_rate is None else str(project.completion_rate),
        'start_date': project.start_date.isoformat() if project.start_date else '',
        'dataPath': f'database://project/{project.id}',
        'weldFile': _data_source_payload(weld_source, '焊口初始化数据.xlsx') if weld_source else None,
        'prefabWeldFile': _data_source_payload(prefab_source, '预制焊口库.xlsx') if prefab_source else None,
        'parserModels': _project_parser_models(project),
        **_project_data_status(project),
        'createdAt': project.created_at.isoformat() if project.created_at else '',
        'updatedAt': project.updated_at.isoformat() if project.updated_at else '',
    }


def _project_values(payload):
    return {
        'project_name': str(payload.get('project_name', '') or '').strip(),
        'pipe_segment': str(payload.get('pipe_segment', '') or '').strip(),
        'prefab_weld_count': _clean_project_int(payload.get('prefab_weld_count')),
        'completion_rate': _clean_project_decimal(payload.get('completion_rate')),
        'start_date': _clean_project_date(payload.get('start_date')),
    }


def _clean_project_decimal(value):
    if value is None or value == '':
        return None
    try:
        return Decimal(str(value).strip())
    except (InvalidOperation, ValueError):
        return value


def _clean_project_int(value):
    if value is None or value == '':
        return 0
    try:
        return int(float(str(value).strip()))
    except (TypeError, ValueError):
        return value


def _clean_project_date(value):
    if value is None or value == '':
        return None
    text = str(value).strip()
    try:
        return datetime.strptime(text[:10], '%Y-%m-%d').date()
    except ValueError:
        return value


def _validate_project_values(values):
    if not values.get('project_name'):
        return '项目名称不能为空'
    if values.get('completion_rate') is not None and not isinstance(values.get('completion_rate'), Decimal):
        return '完成率必须是数字'
    if not isinstance(values.get('prefab_weld_count'), int):
        return '可预制焊口数必须是整数'
    if values.get('start_date') is not None and not hasattr(values.get('start_date'), 'isoformat'):
        return '开始日期格式应为 YYYY-MM-DD'
    return ''


def _safe_project_dir_name(project):
    raw_name = project.project_name
    cleaned = re.sub(r'[<>:"/\\|?*]+', '_', str(raw_name)).strip().strip('.')
    return cleaned


def _project_root(project):
    return PROJECT_DATA_ROOT / _safe_project_dir_name(project)


def _data_source_payload(source, default_name=''):
    if source is None:
        return None
    return {
        'name': source.display_name or default_name,
        'path': source.relative_path,
        'size': source.file_size,
        'updatedAt': source.file_updated_at,
    }


def _project_data_status(project):
    weld_source = _latest_source(project, 'initialization', 'welds') if project else None
    return {
        'hasInitializationData': weld_source is not None,
        'needsInitializationData': weld_source is None,
        'initializationHint': '' if weld_source else '当前项目暂无初始化数据，请先进行文件解析或上传焊口初始化数据。',
    }


def _project_parser_models(project):
    return {
        'idf': _project_parser_model_payload(project, 'idf', 'database'),
        'pcf': _project_parser_model_payload(project, 'pcf', ''),
    }


def _project_parser_model_payload(project, file_type, model_name):
    from pipecloud.models import ParserJob

    job = (
        ParserJob.objects
        .filter(project=project, file_type=file_type)
        .order_by('-updated_at', '-id')
        .first()
    )
    if not job:
        return {
            'fileType': file_type,
            'exists': False,
            'status': 'missing',
            'jobId': '',
            'message': '',
            'modelFile': None,
            'inputFileCount': 0,
            'resultCount': 0,
            'updatedAt': '',
        }

    result_count = len(job.results or [])
    input_file_count = len(job.input_files or [])
    model_file = _parser_job_model_file(job, model_name) if model_name else None
    return {
        'fileType': file_type,
        'exists': bool(model_file) if model_name else job.status == 'completed' and result_count > 0,
        'status': job.status,
        'jobId': job.job_id,
        'message': job.message,
        'current': job.current,
        'percent': job.percent,
        'total': job.total,
        'completed': job.completed,
        'failed': job.failed,
        'modelFile': model_file,
        'inputFileCount': input_file_count,
        'resultCount': result_count,
        'updatedAt': job.updated_at.isoformat(timespec='seconds') if job.updated_at else '',
    }


def _parser_job_model_file(job, model_name):
    if model_name != 'database':
        return None
    from pipecloud.services.idf_model_storage import idf_database_file_payload

    try:
        model = job.idf_model
    except ObjectDoesNotExist:
        return None
    if model.status != 'ready':
        return None
    return idf_database_file_payload(model)


def _safe_upload_name(name):
    file_name = Path(str(name or '').replace('\\', '/').split('/')[-1]).name
    return file_name or f'upload-{datetime.now().strftime("%H%M%S")}'


def _parser_file_type(file_name):
    suffix = Path(file_name).suffix.lower()
    if suffix == '.idf':
        return 'idf'
    if suffix == '.pcf':
        return 'pcf'
    return ''


def _parser_download_url(file_path):
    relative_path = Path(file_path).resolve().relative_to(FILE_PARSER_ROOT.resolve())
    relative_text = str(relative_path).replace('\\', '/')
    return f'/api/pipecloud/file-parser/download/?path={quote(relative_text)}'


def _parser_relative_path(file_path):
    return str(Path(file_path).resolve().relative_to(FILE_PARSER_ROOT.resolve())).replace('\\', '/')


def _resolve_parser_file(relative_path):
    target_path = (FILE_PARSER_ROOT / (relative_path or '')).resolve()
    target_path.relative_to(FILE_PARSER_ROOT.resolve())
    if not target_path.exists() or not target_path.is_file():
        raise FileNotFoundError('暂存文件不存在')
    return target_path


def _parse_project_payload(request):
    try:
        payload = json.loads(request.body.decode('utf-8') or '{}')
    except (UnicodeDecodeError, json.JSONDecodeError):
        return None, '请求内容不是有效 JSON'
    if not isinstance(payload, dict):
        return None, '项目数据无效'
    values = _project_values(payload)
    error = _validate_project_values(values)
    if error:
        return None, error
    return values, None


def _project_file_response():
    rows = []
    for project in Project.objects.all():
        rows.append({
            column['title']: getattr(project, column['field'])
            for column in PROJECT_COLUMNS
        })

    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        pd.DataFrame(rows, columns=[column['title'] for column in PROJECT_COLUMNS]).to_excel(
            writer,
            index=False,
            sheet_name='项目',
        )

    output.seek(0)
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="projects.xlsx"'
    return response


def _project_or_error(project_id):
    try:
        return Project.objects.get(pk=project_id), ''
    except Project.DoesNotExist:
        return None, '项目不存在'


def _is_date_dir(path):
    return path.is_dir() and len(path.name) == 8 and path.name.isdigit()


def _is_date_value(value):
    return len(str(value)) == 8 and str(value).isdigit()


def _today_plan_date():
    return datetime.now().strftime('%Y%m%d')


def _plan_folder_date(folder_path):
    if _is_date_value(folder_path.name):
        return folder_path.name
    return datetime.fromtimestamp(folder_path.stat().st_mtime).strftime('%Y%m%d')


def _plan_record_payload(record):
    files = record.files or []
    return {
        'id': record.id,
        'planKey': record.plan_key,
        'planName': record.plan_name,
        'planDate': record.plan_date,
        'planFolder': record.plan_folder,
        'name': record.plan_folder,
        'path': record.relative_path,
        'fileCount': len(files) if files else record.file_count,
        'updatedAt': record.folder_updated_at,
        'files': files,
        'summary': record.summary or {},
    }


def _sync_plan_records(project, plan_sources):
    ensure_project_tables(project)
    return None


def _plan_records_for_source(project, source):
    ensure_project_tables(project)
    with using_project_tables(project):
        return list(PlanRecord.objects.filter(project=project, plan_key=source['key']))


def _plan_records_by_date(records, selected_date):
    return [_plan_record_payload(record) for record in records if record.plan_date == selected_date]


def _plan_group_payload(records, today_date):
    today_records = sorted(
        [record for record in records if record.plan_date == today_date],
        key=lambda record: (-record.folder_updated_at, record.plan_folder),
    )
    future_records = sorted(
        [record for record in records if record.plan_date > today_date],
        key=lambda record: (record.plan_date, record.plan_folder),
    )
    history_records = sorted(
        [record for record in records if record.plan_date < today_date],
        key=lambda record: (record.plan_date, record.folder_updated_at),
        reverse=True,
    )
    return {
        'todayPlans': [_plan_record_payload(record) for record in today_records],
        'futurePlans': [_plan_record_payload(record) for record in future_records],
        'historyPlans': [_plan_record_payload(record) for record in history_records],
        'todayCount': len(today_records),
        'futureCount': len(future_records),
        'historyCount': len(history_records),
    }


def _plan_source_payload(project, source, selected_date=None):
    records = _plan_records_for_source(project, source)
    dates = sorted({record.plan_date for record in records}, reverse=True)
    selected_date = selected_date if selected_date in dates else (dates[0] if dates else '')
    plans = _plan_records_by_date(records, selected_date)

    return {
        'key': source['key'],
        'name': source['name'],
        'root': f'database://plan/{source["key"]}',
        'exists': bool(records),
        'dates': dates,
        'selectedDate': selected_date,
        'selectedPath': f'database://plan/{source["key"]}/{selected_date}' if selected_date else '',
        'plans': plans,
        'files': plans,
        'fileCount': len(plans),
        **_plan_group_payload(records, _today_plan_date()),
    }


def _plan_date_file_count(project, source, date):
    ensure_project_tables(project)
    with using_project_tables(project):
        return PlanRecord.objects.filter(project=project, plan_key=source['key'], plan_date=date).count()


def _resolve_plan_file(plan_key, plan_folder, file_name, plan_sources):
    source = plan_sources.get(plan_key)
    if source is None:
        return None, '未知计划类型'
    if not plan_folder or Path(plan_folder).name != plan_folder:
        return None, '计划文件夹无效'
    if not file_name or Path(file_name).name != file_name:
        return None, '文件名无效'

    return {
        'plan_key': plan_key,
        'plan_folder': plan_folder,
        'file_name': file_name,
        'source_key': f'{plan_key}:{plan_folder}:{file_name}',
    }, ''


def _plan_payload(project, plan_key, selected_date=None, plan_sources=None):
    plan_sources = plan_sources or PLAN_SOURCES
    if plan_key == 'all':
        sources = list(plan_sources.values())
    else:
        source = plan_sources.get(plan_key)
        if source is None:
            return None
        sources = [source]

    _sync_plan_records(project, plan_sources)
    source_payloads = [_plan_source_payload(project, source, selected_date) for source in sources]
    all_dates = sorted({date for source in source_payloads for date in source['dates']}, reverse=True)
    records_by_source_date = {
        source_payload['key']: {
            date: _plan_records_by_date(_plan_records_for_source(project, plan_sources[source_payload['key']]), date)
            for date in source_payload['dates']
        }
        for source_payload in source_payloads
    }
    timeline = [
        {
            'date': date,
            'plans': [
                {
                    'planKey': source['key'],
                    'planName': source['name'],
                    'fileCount': _plan_date_file_count(project, plan_sources[source['key']], date),
                    'available': date in source['dates'],
                    'records': records_by_source_date.get(source['key'], {}).get(date, []),
                }
                for source in source_payloads
            ],
        }
        for date in all_dates
    ]

    return {
        'key': plan_key,
        'name': '排产计划' if plan_key == 'all' else source_payloads[0]['name'],
        'selectedDate': selected_date or (all_dates[0] if all_dates else ''),
        'dates': all_dates,
        'sources': source_payloads,
        'timeline': timeline,
    }


def _source_payload_from_database(source, total=0, unit_column='单元号'):
    return {
        'path': source.relative_path if source else '',
        'exists': bool(source) or total > 0,
        'total': int(total),
        'unitColumn': unit_column if total else '',
        'units': {},
    }


def _unit_counts_from_queryset(queryset, source=None):
    units = {}
    total = 0
    for row in queryset.values('unit').annotate(count=Count('id')):
        unit = str(row.get('unit') or '').strip() or '未分单元'
        count = int(row.get('count') or 0)
        units[unit] = units.get(unit, 0) + count
        total += count
    payload = _source_payload_from_database(source, total)
    payload['units'] = units
    return payload


def _latest_source(project, source_type, source_key=None, display_name=None):
    ensure_project_tables(project)
    with using_project_tables(project):
        queryset = DataSourceFile.objects.filter(project=project, source_type=source_type)
        if source_key is not None:
            queryset = queryset.filter(source_key=source_key)
        if display_name is not None:
            queryset = queryset.filter(display_name=display_name)
        return queryset.order_by('-file_updated_at', '-id').first()


def _plan_primary_file_name(record):
    files = record.files or []
    primary_name = WELDING_PRIMARY_PLAN_FILE_NAME
    primary_stem = Path(primary_name).stem
    file_info = next((item for item in files if item.get('name') == primary_name), None)
    if file_info is None:
        file_info = next((item for item in files if str(item.get('name', '')).startswith(primary_stem)), None)
    if file_info is None:
        return ''
    return Path(file_info.get('name', '')).name


def _welding_dashboard_payload(project, data_root, force_refresh=False):
    source = _plan_sources(data_root)['welding']
    records = _plan_records_for_source(project, source)
    today_date = _today_plan_date()
    today_records = [record for record in records if record.plan_date == today_date]
    history_records = [record for record in records if record.plan_date != today_date]

    with using_project_tables(project):
        source_by_key = {
            source_file.source_key: source_file
            for source_file in DataSourceFile.objects.filter(
                project=project,
                source_type='plan',
                source_key__startswith='welding:',
            )
        }
        rows = []
        unique_welds = {}
        unique_history_welds = {}
        unique_today_welds = {}
        for record in records:
            file_name = _plan_primary_file_name(record)
            source_file = source_by_key.get(f'welding:{record.plan_folder}:{file_name}') if file_name else None
            queryset = WeldingPlanRow.objects.filter(project=project, source_file=source_file) if source_file else WeldingPlanRow.objects.none()
            total_rows = queryset.count()
            completed_rows = sum(1 for value in queryset.values_list('completed_flag', flat=True) if _truthy_text(value))
            for plan_row in queryset.values(
                'weld_no_final',
                'weld_no_start',
                'pipeline',
                'unit',
                'diameter',
                'completed_flag',
            ):
                key = '|'.join(
                    str(plan_row.get(field) or '').strip()
                    for field in ('weld_no_final', 'weld_no_start', 'pipeline', 'unit', 'diameter')
                )
                completed = _truthy_text(plan_row.get('completed_flag'))
                unique_welds[key] = unique_welds.get(key, False) or completed
                target = unique_today_welds if record.plan_date == today_date else unique_history_welds
                target[key] = target.get(key, False) or completed
            rows.append({
                'planDate': record.plan_date,
                'planFolder': record.plan_folder,
                'fileName': file_name,
                'path': source_file.relative_path if source_file else record.relative_path,
                'totalRows': total_rows,
                'completedRows': completed_rows,
                'completionRate': _percentage(completed_rows, total_rows),
                'updatedAt': source_file.file_updated_at if source_file else record.folder_updated_at,
            })

    history_rows = [row for row in rows if row['planDate'] != today_date]
    today_rows = [row for row in rows if row['planDate'] == today_date]

    total_rows = len(unique_welds)
    completed_rows = sum(unique_welds.values())
    history_total_rows = len(unique_history_welds)
    history_completed_rows = sum(unique_history_welds.values())
    today_total_rows = len(unique_today_welds)
    today_completed_rows = sum(unique_today_welds.values())

    payload = {
        'projectId': project.id,
        'projectName': project.project_name,
        'updatedAt': datetime.now().isoformat(timespec='seconds'),
        'todayDate': today_date,
        'planCount': len(records),
        'historyPlanCount': len(history_records),
        'todayPlanCount': len(today_records),
        'totalRows': total_rows,
        'completedRows': completed_rows,
        'completionRate': _percentage(completed_rows, total_rows),
        'historyTotalRows': history_total_rows,
        'historyCompletedRows': history_completed_rows,
        'historyCompletionRate': _percentage(history_completed_rows, history_total_rows),
        'todayTotalRows': today_total_rows,
        'todayCompletedRows': today_completed_rows,
        'todayCompletionRate': _percentage(today_completed_rows, today_total_rows),
        'recentPlans': rows[:8],
    }
    return payload


def _dashboard_quantity(value):
    if value in (None, ''):
        return Decimal('0')
    text = str(value).strip().replace(',', '')
    if not text or text.lower() == 'nan':
        return Decimal('0')
    try:
        return Decimal(text)
    except (InvalidOperation, ValueError):
        return Decimal('0')


def _dashboard_number(value):
    number = float(value)
    return round(number, 3)


def _dashboard_percentage(numerator, denominator):
    if denominator <= 0:
        return 0
    return round(float(min(numerator / denominator, Decimal('1')) * 100), 2)


def _arrival_date_text(*values):
    for value in values:
        text = str(value or '').strip()
        if not text:
            continue
        match = re.search(r'(\d{4})[-/.年]?(\d{1,2})[-/.月]?(\d{1,2})', text)
        if match:
            year, month, day = match.groups()
            return f'{year}-{month.zfill(2)}-{day.zfill(2)}'
        match = re.search(r'(\d{4})(\d{2})(\d{2})', text)
        if match:
            year, month, day = match.groups()
            return f'{year}-{month}-{day}'
        parsed = pd.to_datetime(text, errors='coerce')
        if pd.notna(parsed):
            return parsed.strftime('%Y-%m-%d')
    return '未识别日期'


def _arrival_material_type(row):
    unit = str(row.get('unit') or '').strip()
    name = str(row.get('name') or '').strip()
    category = str(row.get('material_category') or '').strip()
    pipe_categories = {'管材', '直管', '管子', '钢管'}
    return 'pipe' if unit == '根' or name == '管子' or category in pipe_categories else 'other'


def _arrival_material_qty(row):
    actual = _dashboard_quantity(row.get('actual_arrival_qty'))
    if actual > 0:
        return actual
    return _dashboard_quantity(row.get('shipment_qty'))


def _collect_stock_materials(project, model, source_keys):
    materials = {}
    for row in model.objects.filter(project=project, source_file__source_key__in=source_keys).values(
        'material_code',
        'stock_qty',
        'material_description',
        'unit',
    ):
        code = str(row.get('material_code') or '').strip()
        if not code:
            continue
        item = materials.setdefault(code, {
            'actualQty': Decimal('0'),
            'description': '',
            'unit': '',
        })
        item['actualQty'] += _dashboard_quantity(row.get('stock_qty'))
        item['description'] = item['description'] or str(row.get('material_description') or '').strip()
        item['unit'] = str(row.get('unit') or '').strip() or item['unit']
    return materials


def _arrival_date_stats(project):
    ensure_project_tables(project)
    with using_project_tables(project):
        source_dates = {
            row['source_file_id']: _arrival_date_text(
                row.get('arrival_time'),
                row.get('shipment_date'),
                row.get('source_file__display_name'),
                row.get('source_file__file_updated_at'),
            )
            for row in ArrivalOrderRow.objects.filter(project=project).values(
                'source_file_id',
                'arrival_time',
                'shipment_date',
                'source_file__display_name',
                'source_file__file_updated_at',
            )
        }
        stats = {}
        for row in ArrivalMaterialRow.objects.filter(project=project).values(
            'source_file_id',
            'source_file__display_name',
            'source_file__file_updated_at',
            'actual_arrival_qty',
            'shipment_qty',
            'actual_arrival_count',
            'unit',
            'name',
            'material_category',
        ):
            arrival_date = source_dates.get(row['source_file_id']) or _arrival_date_text(
                row.get('source_file__display_name'),
                row.get('source_file__file_updated_at'),
            )
            material_type = _arrival_material_type(row)
            item = stats.setdefault((arrival_date, material_type), {
                'date': arrival_date,
                'materialType': material_type,
                'quantity': Decimal('0'),
                'rowCount': 0,
                'pipeCount': Decimal('0'),
            })
            item['quantity'] += _arrival_material_qty(row)
            item['rowCount'] += 1
            if material_type == 'pipe':
                item['pipeCount'] += _dashboard_quantity(row.get('actual_arrival_count'))

    rows = []
    for item in stats.values():
        rows.append({
            'date': item['date'],
            'materialType': item['materialType'],
            'quantity': _dashboard_number(item['quantity']),
            'rowCount': item['rowCount'],
            'pipeCount': _dashboard_number(item['pipeCount']),
        })
    rows.sort(key=lambda item: (item['date'] == '未识别日期', item['date'], item['materialType'] != 'pipe'))
    return rows


def _arrival_material_dashboard_payload(project):
    """Compare required weld-library quantities with raw arrival quantities."""
    ensure_project_tables(project)
    grouped = {}

    def ensure_row(material_type, code):
        key = (material_type, code)
        if key not in grouped:
            grouped[key] = {
                'materialType': material_type,
                'materialCode': code,
                'description': '',
                'unit': '米' if material_type == 'pipe' else '',
                'expectedQty': Decimal('0'),
                'actualQty': Decimal('0'),
                'requiredActualQty': Decimal('0'),
                'extraQty': Decimal('0'),
            }
        return grouped[key]

    with using_project_tables(project):
        pipe_materials = _collect_stock_materials(project, PipeMaterialRow, [
            'pipe-library',
            'anti-pipe-library',
        ])
        fitting_materials = _collect_stock_materials(project, FittingMaterialRow, [
            'fitting-library',
            'anti-fitting-library',
        ])

        pipe_codes = set(pipe_materials)
        fitting_codes = set(fitting_materials)
        for row in WeldLibraryRow.objects.filter(project=project).values(
            'material_code_1', 'material_code_2',
            'material_mark_1', 'material_mark_2',
            'quantity_1', 'quantity_2',
            'description_1', 'description_2',
        ):
            for side in (1, 2):
                code = str(row.get(f'material_code_{side}') or '').strip()
                if not code:
                    continue
                mark = str(row.get(f'material_mark_{side}') or '').strip().upper()
                if mark == 'P' or (not mark and code in pipe_codes and code not in fitting_codes):
                    material_type = 'pipe'
                else:
                    material_type = 'other'
                item = ensure_row(material_type, code)
                item['expectedQty'] += _dashboard_quantity(row.get(f'quantity_{side}'))
                item['description'] = item['description'] or str(row.get(f'description_{side}') or '').strip()

        for (material_type, code), item in grouped.items():
            library_item = (pipe_materials if material_type == 'pipe' else fitting_materials).get(code)
            if library_item is None:
                continue
            item['actualQty'] = library_item['actualQty']
            item['description'] = library_item['description'] or item['description']
            item['unit'] = library_item['unit'] or item['unit']

    rows = []
    for item in grouped.values():
        expected = item.pop('expectedQty')
        actual = item.pop('actualQty')
        required_actual = min(actual, expected)
        extra = max(actual - expected, Decimal('0'))
        difference = expected - required_actual
        rows.append({
            **item,
            'expectedQty': _dashboard_number(expected),
            'actualQty': _dashboard_number(actual),
            'requiredActualQty': _dashboard_number(required_actual),
            'extraQty': _dashboard_number(extra),
            'differenceQty': _dashboard_number(difference),
            'arrivalRate': _dashboard_percentage(required_actual, expected),
        })
    rows.sort(key=lambda item: (item['materialType'] != 'pipe', item['materialCode']))

    summaries = {}
    for material_type in ('pipe', 'other'):
        category_rows = [row for row in rows if row['materialType'] == material_type]
        expected = sum((_dashboard_quantity(row['expectedQty']) for row in category_rows), Decimal('0'))
        actual = sum((_dashboard_quantity(row['actualQty']) for row in category_rows), Decimal('0'))
        required_actual = sum((_dashboard_quantity(row['requiredActualQty']) for row in category_rows), Decimal('0'))
        extra = sum((_dashboard_quantity(row['extraQty']) for row in category_rows), Decimal('0'))
        summaries[material_type] = {
            'expectedQty': _dashboard_number(expected),
            'actualQty': _dashboard_number(actual),
            'requiredActualQty': _dashboard_number(required_actual),
            'extraQty': _dashboard_number(extra),
            'differenceQty': _dashboard_number(expected - required_actual),
            'arrivalRate': _dashboard_percentage(required_actual, expected),
            'materialCount': len(category_rows),
        }

    return {
        'projectId': project.id,
        'projectName': project.project_name,
        'updatedAt': datetime.now().isoformat(timespec='seconds'),
        'summaries': summaries,
        'rows': rows,
        'dateStats': _arrival_date_stats(project),
    }


def _plan_summary_number(summary, key):
    return _dashboard_number(_dashboard_quantity((summary or {}).get(key)))


def _plan_summary_list(summary, key):
    value = (summary or {}).get(key) or []
    if isinstance(value, str):
        return [item.strip() for item in value.replace(',', '、').split('、') if item.strip()]
    if isinstance(value, list):
        return [str(item or '').strip() for item in value if str(item or '').strip()]
    return []


def _pre_schedule_dashboard_counts(project, source_key):
    source = _latest_source(project, 'pre-schedule', source_key)
    if source is None:
        return {
            'path': '',
            'totalRows': 0,
            'schedulableRows': 0,
            'rejectedRows': 0,
            'statusRows': [],
        }

    selected_sheet, _, _, _, rows = table_payload(source, PRE_SCHEDULE_MODELS, '预排产匹配结果')
    status_counts = {}
    for row in rows:
        status = str(row.get('预排产状态') or '').strip() or '未标记'
        status_counts[status] = status_counts.get(status, 0) + 1
    schedulable = status_counts.get('可预排产', 0)
    total = len(rows)
    return {
        'path': source.relative_path,
        'sheet': selected_sheet,
        'totalRows': total,
        'schedulableRows': schedulable,
        'rejectedRows': max(total - schedulable, 0),
        'statusRows': [
            {'status': status, 'count': count}
            for status, count in sorted(status_counts.items(), key=lambda item: item[0])
        ],
    }


def _anti_corrosion_dashboard_payload(project, data_root):
    source = _plan_sources(data_root)['anti-corrosion']
    records = _plan_records_for_source(project, source)
    today_date = _today_plan_date()
    pre_schedule = _pre_schedule_dashboard_counts(project, 'anti-corrosion-pre-schedule')

    with using_project_tables(project):
        commission_rows = list(
            MasterScheduleRow.objects
            .filter(project=project)
            .exclude(anti_corrosion_date='')
            .values(
                'anti_corrosion_order_no',
                'anti_corrosion_date',
                'library_seq',
                'unit',
                'pipeline',
                'segment_no',
                'diameter',
                'stage_payload',
                'updated_at',
            )
        )

    grouped = {}
    for row in commission_rows:
        payload = dict((row.get('stage_payload') or {}).get('anti-corrosion') or {})
        commission_no = str(payload.get('防腐委托单号') or row.get('anti_corrosion_order_no') or '').strip() or '未编号'
        item = grouped.setdefault(commission_no, {
            'commissionNo': commission_no,
            'commissionDate': str(payload.get('委托日期') or row.get('anti_corrosion_date') or '').strip(),
            'weldCount': 0,
            'segmentKeys': set(),
            'totalArea': Decimal('0'),
            'diameterTotal': Decimal('0'),
            'units': set(),
            'pipelines': set(),
            'updatedAt': row.get('updated_at').timestamp() if row.get('updated_at') else 0,
        })
        item['weldCount'] += 1
        item['totalArea'] += _dashboard_quantity(payload.get('防腐面积'))
        item['diameterTotal'] += _dashboard_quantity(payload.get('寸径') or row.get('diameter'))
        segment_key = tuple(
            str(payload.get(column) or row.get(field) or '').strip()
            for field, column in (('unit', '单元号'), ('pipeline', '管线号'), ('segment_no', '管段号'))
        )
        if any(segment_key):
            item['segmentKeys'].add(segment_key)
        for key, target in (('unit', 'units'), ('pipeline', 'pipelines')):
            value = str(payload.get({'unit': '单元号', 'pipeline': '管线号'}[key]) or row.get(key) or '').strip()
            if value:
                item[target].add(value)
        item['updatedAt'] = max(float(item['updatedAt'] or 0), row.get('updated_at').timestamp() if row.get('updated_at') else 0)

    rows = []
    for item in grouped.values():
        rows.append({
            'commissionNo': item['commissionNo'],
            'commissionDate': item['commissionDate'],
            'weldCount': item['weldCount'],
            'segmentCount': len(item['segmentKeys']),
            'totalArea': _dashboard_number(item['totalArea']),
            'diameterTotal': _dashboard_number(item['diameterTotal']),
            'unitCount': len(item['units']),
            'pipelineCount': len(item['pipelines']),
            'updatedAt': item['updatedAt'],
        })
    rows.sort(key=lambda item: (item['commissionDate'], item['commissionNo']), reverse=True)

    plan_rows = [
        {
            'planDate': record.plan_date,
            'planFolder': record.plan_folder,
            'fileCount': record.file_count,
            'updatedAt': record.folder_updated_at,
        }
        for record in records[:8]
    ]

    total_area = sum((_dashboard_quantity(row['totalArea']) for row in rows), Decimal('0'))
    segment_keys = {
        (
            str(row.get('unit') or '').strip(),
            str(row.get('pipeline') or '').strip(),
            str(row.get('segment_no') or '').strip(),
        )
        for row in commission_rows
        if any((str(row.get('unit') or '').strip(), str(row.get('pipeline') or '').strip(), str(row.get('segment_no') or '').strip()))
    }
    today_records = [record for record in records if record.plan_date == today_date]
    return {
        'projectId': project.id,
        'projectName': project.project_name,
        'updatedAt': datetime.now().isoformat(timespec='seconds'),
        'todayDate': today_date,
        'planCount': len(records),
        'todayPlanCount': len(today_records),
        'commissionCount': len(rows),
        'weldCount': len(commission_rows),
        'segmentCount': len(segment_keys),
        'totalArea': _dashboard_number(total_area),
        'preSchedule': pre_schedule,
        'rows': rows[:200],
        'recentPlans': plan_rows,
    }


def _cutting_dashboard_payload(project, data_root):
    source = _plan_sources(data_root)['cutting']
    records = _plan_records_for_source(project, source)
    today_date = _today_plan_date()
    pre_schedule = _pre_schedule_dashboard_counts(project, 'weld-pre-schedule')

    fallback_by_date = {}
    if any(not (record.summary or {}).get('weldCount') for record in records):
        ensure_project_tables(project)
        with using_project_tables(project):
            plan_rows = list(
                WeldingPlanRow.objects
                .filter(project=project)
                .exclude(cut_date='')
                .values('cut_date', 'cut_order_no', 'weld_order_no', 'diameter')
            )
        for plan_row in plan_rows:
            item = fallback_by_date.setdefault(str(plan_row.get('cut_date') or ''), {
                'orderNumbers': [],
                'relatedOrderNumbers': [],
                'weldCount': 0,
                'diameterTotal': Decimal('0'),
            })
            cut_order_no = str(plan_row.get('cut_order_no') or '').strip()
            weld_order_no = str(plan_row.get('weld_order_no') or '').strip()
            if cut_order_no and cut_order_no not in item['orderNumbers']:
                item['orderNumbers'].append(cut_order_no)
            if weld_order_no and weld_order_no not in item['relatedOrderNumbers']:
                item['relatedOrderNumbers'].append(weld_order_no)
            item['weldCount'] += 1
            item['diameterTotal'] += _dashboard_quantity(plan_row.get('diameter'))

    rows = []
    for record in records:
        summary = record.summary or {}
        fallback = fallback_by_date.get(record.plan_date, {})
        order_numbers = _plan_summary_list(summary, 'orderNumbers') or fallback.get('orderNumbers', [])
        related_order_numbers = _plan_summary_list(summary, 'relatedOrderNumbers') or fallback.get('relatedOrderNumbers', [])
        rows.append({
            'planDate': record.plan_date,
            'planFolder': record.plan_folder,
            'fileCount': record.file_count,
            'orderCount': int(summary.get('orderCount') or len(order_numbers) or 0),
            'weldCount': int(_dashboard_quantity(summary.get('weldCount') or fallback.get('weldCount'))),
            'diameterTotal': _plan_summary_number(summary, 'diameterTotal') or _dashboard_number(fallback.get('diameterTotal', 0)),
            'orderNumbers': order_numbers,
            'orderNumbersText': '、'.join(order_numbers),
            'relatedOrderNumbers': related_order_numbers,
            'relatedOrderNumbersText': '、'.join(related_order_numbers),
            'updatedAt': record.folder_updated_at,
        })
    rows.sort(key=lambda item: (item['planDate'], item['updatedAt']), reverse=True)

    today_rows = [row for row in rows if row['planDate'] == today_date]
    return {
        'projectId': project.id,
        'projectName': project.project_name,
        'updatedAt': datetime.now().isoformat(timespec='seconds'),
        'todayDate': today_date,
        'planCount': len(records),
        'todayPlanCount': len(today_rows),
        'orderCount': sum(row['orderCount'] for row in rows),
        'todayOrderCount': sum(row['orderCount'] for row in today_rows),
        'weldCount': sum(row['weldCount'] for row in rows),
        'todayWeldCount': sum(row['weldCount'] for row in today_rows),
        'diameterTotal': _dashboard_number(sum((_dashboard_quantity(row['diameterTotal']) for row in rows), Decimal('0'))),
        'preSchedule': pre_schedule,
        'rows': rows[:200],
    }


def _number_value(value, default=0.0):
    number = pd.to_numeric(value, errors='coerce')
    return float(number) if pd.notna(number) else default


def _parse_number_list(value):
    if isinstance(value, list):
        return [_number_value(item) for item in value if pd.notna(item)]

    if pd.isna(value) or str(value).strip() == '':
        return []

    text = str(value).strip()
    try:
        parsed = json.loads(text)
        if isinstance(parsed, list):
            return [_number_value(item) for item in parsed if pd.notna(item)]
    except json.JSONDecodeError:
        pass

    numbers = []
    for part in text.replace('，', ',').replace(';', ',').replace('；', ',').split(','):
        part = part.strip()
        if part:
            numbers.append(_number_value(part))
    return numbers


def _text_value(row, column_name):
    if column_name not in row.index:
        return ''
    value = row.get(column_name)
    return '' if pd.isna(value) else str(value).strip()


def _cutting_pipe_payload(row):
    original_length = _number_value(row.get(CUTTING_COLUMNS['original_length']))
    if original_length <= 0:
        original_length = _number_value(row.get(CUTTING_COLUMNS['stock_qty']))
    remaining_length = _number_value(row.get(CUTTING_COLUMNS['remaining_length']))
    cut_lengths = _parse_number_list(row.get(CUTTING_COLUMNS['cut_lengths']))
    loss_lengths = _parse_number_list(row.get(CUTTING_COLUMNS['loss_lengths']))
    consumed_lengths = _parse_number_list(row.get(CUTTING_COLUMNS['consumed_lengths']))

    if len(consumed_lengths) != len(cut_lengths):
        consumed_lengths = [
            round(cut + (loss_lengths[index] if index < len(loss_lengths) else 0), 3)
            for index, cut in enumerate(cut_lengths)
        ]

    used_length = round(sum(consumed_lengths), 3)
    if remaining_length <= 0 and original_length > used_length:
        remaining_length = round(original_length - used_length, 3)

    segments = []
    cursor = 0.0
    for index, cut_length in enumerate(cut_lengths):
        consumed_length = consumed_lengths[index] if index < len(consumed_lengths) else cut_length
        loss_length = loss_lengths[index] if index < len(loss_lengths) else max(consumed_length - cut_length, 0)
        segment_length = max(float(consumed_length), 0)
        if segment_length <= 0:
            continue
        segments.append({
            'type': 'cut',
            'index': index + 1,
            'label': round(float(cut_length), 3),
            'length': round(float(cut_length), 3),
            'consumedLength': round(segment_length, 3),
            'lossLength': round(float(loss_length), 3),
            'start': round(cursor, 3),
            'percent': round(segment_length / original_length * 100, 4) if original_length > 0 else 0,
        })
        cursor = round(cursor + segment_length, 3)

    if remaining_length > 0:
        segments.append({
            'type': 'remaining',
            'index': None,
            'label': round(remaining_length, 3),
            'length': round(remaining_length, 3),
            'consumedLength': round(remaining_length, 3),
            'lossLength': 0,
            'start': round(cursor, 3),
            'percent': round(remaining_length / original_length * 100, 4) if original_length > 0 else 0,
        })

    return {
        'pipeNo': _text_value(row, CUTTING_COLUMNS['pipe_no']),
        'inventoryType': _text_value(row, '库存类型'),
        'materialCode': _text_value(row, CUTTING_COLUMNS['material_code']),
        'description': _text_value(row, CUTTING_COLUMNS['description']),
        'heatNo': _text_value(row, CUTTING_COLUMNS['heat_no']),
        'originalLength': round(original_length, 3),
        'usedLength': used_length,
        'remainingLength': round(max(remaining_length, 0), 3),
        'utilization': round(used_length / original_length * 100, 1) if original_length > 0 else 0,
        'cutCount': len(cut_lengths),
        'segments': segments,
    }


def _row_payload(columns, values):
    padded_values = list(values or [])[:len(columns)]
    if len(padded_values) < len(columns):
        padded_values.extend([''] * (len(columns) - len(padded_values)))
    return {
        column: _clean_cell(value)
        for column, value in zip(columns, padded_values)
    }


def _read_excel_rows(file_path, sheet_name):
    workbook = load_workbook(file_path, read_only=True, data_only=True)
    try:
        sheets = workbook.sheetnames
        selected_sheet = sheet_name if sheet_name in sheets else sheets[0]
        worksheet = workbook[selected_sheet]

        rows_iter = worksheet.iter_rows(values_only=True)
        header = next(rows_iter, [])
        columns = [str(column or '') for column in header]
        rows = [_row_payload(columns, values) for values in rows_iter]
        return selected_sheet, sheets, len(rows), columns, rows
    finally:
        workbook.close()


def _percentage(numerator, denominator):
    denominator = int(denominator or 0)
    if denominator <= 0:
        return 0
    return round(int(numerator or 0) / denominator * 100, 2)


def _initialization_stats_payload(project, data_root, force_refresh=False):
    with using_project_tables(project):
        total_source = _latest_source(project, 'initialization', 'welds')
        library_source = _latest_source(project, 'library', 'weld-library')
        total_counts = _unit_counts_from_queryset(
            InitializationWeldRow.objects.filter(project=project),
            total_source,
        )
        prefab_counts = _unit_counts_from_queryset(
            WeldLibraryRow.objects.filter(project=project),
            library_source,
        )
        auto_counts = _unit_counts_from_queryset(
            WeldLibraryRow.objects.filter(project=project, welding_mode__contains='自动焊'),
            library_source,
        )

    units = sorted(set(total_counts['units']) | set(prefab_counts['units']) | set(auto_counts['units']))

    unit_rows = [
        {
            'unit': unit,
            'totalWeldCount': int(total_counts['units'].get(unit, 0)),
            'prefabWeldCount': int(prefab_counts['units'].get(unit, 0)),
            'autoWeldCount': int(auto_counts['units'].get(unit, 0)),
            'prefabRate': _percentage(prefab_counts['units'].get(unit, 0), total_counts['units'].get(unit, 0)),
            'autoRate': _percentage(auto_counts['units'].get(unit, 0), prefab_counts['units'].get(unit, 0)),
        }
        for unit in units
    ]

    payload = {
        'projectId': project.id,
        'projectName': project.project_name,
        'updatedAt': datetime.now().isoformat(timespec='seconds'),
        'sources': {
            'total': total_counts,
            'prefab': prefab_counts,
            'auto': auto_counts,
        },
        'totalWeldCount': total_counts['total'],
        'prefabWeldCount': prefab_counts['total'],
        'autoWeldCount': auto_counts['total'],
        'prefabRate': _percentage(prefab_counts['total'], total_counts['total']),
        'autoRate': _percentage(auto_counts['total'], prefab_counts['total']),
        'unitCount': len(unit_rows),
        'units': unit_rows,
    }
    return payload


def _truthy_text(value):
    return str(value if value is not None else '').strip().lower() in {
        'true',
        '1',
        'yes',
        'y',
        '完成',
        '已完成',
        'done',
        'finished',
    }


def _resolve_column(columns, canonical_name):
    aliases = [canonical_name, *WELD_COLUMN_ALIASES.get(canonical_name, [])]
    normalized = {str(column).strip(): column for column in columns}
    for alias in aliases:
        if alias in normalized:
            return normalized[alias]
    return ''


def _weld_key_series(dataframe):
    key_columns = [
        WELD_COLUMNS['weld_no_final'],
        WELD_COLUMNS['weld_no_start'],
        WELD_COLUMNS['pipeline'],
        WELD_COLUMNS['unit'],
        WELD_COLUMNS['diameter'],
    ]
    existing = [_resolve_column(dataframe.columns, column) for column in key_columns]
    existing = [column for column in existing if column]
    if not existing:
        seq_col = _resolve_column(dataframe.columns, WELD_COLUMNS['library_seq'])
        if seq_col:
            return dataframe[seq_col].astype('string').fillna('').str.strip()
        return pd.Series(range(len(dataframe)), index=dataframe.index).astype(str)

    key_df = dataframe[existing].copy()
    for column in existing:
        raw_series = key_df[column]
        numeric_series = pd.to_numeric(raw_series, errors='coerce')
        normalized = raw_series.astype('string')
        numeric_mask = numeric_series.notna()
        if numeric_mask.any():
            normalized.loc[numeric_mask] = numeric_series.loc[numeric_mask].map(lambda value: format(float(value), 'g'))
        key_df[column] = normalized.fillna('').str.strip()
    return key_df.agg('|'.join, axis=1)


def _source_dataframe(source, sheet_models, sheet_name=None):
    if source is None:
        return pd.DataFrame(), '', []
    selected_sheet, sheets, _, columns, rows = table_payload(source, sheet_models, sheet_name)
    return pd.DataFrame(rows, columns=columns), selected_sheet, sheets


def _prefab_segment_count(dataframe):
    """Count segments by unit, pipeline and segment number."""
    segment_col = _resolve_column(dataframe.columns, WELD_COLUMNS['segment_no'])
    if not segment_col:
        return 0

    key_columns = []
    for column_name in (WELD_COLUMNS['unit'], WELD_COLUMNS['pipeline'], WELD_COLUMNS['segment_no']):
        resolved = _resolve_column(dataframe.columns, column_name)
        if resolved and resolved not in key_columns:
            key_columns.append(resolved)

    segment_keys = dataframe[key_columns].copy()
    for column in key_columns:
        segment_keys[column] = segment_keys[column].fillna('').astype(str).str.strip()
    segment_keys = segment_keys[segment_keys[segment_col].ne('')]
    return int(len(segment_keys.drop_duplicates(subset=key_columns)))


def _update_project_weld_metrics(
    project,
    data_root,
    update_segment_count=True,
    update_prefab_weld_count=True,
    update_completion_rate=True,
):
    source = _latest_source(project, 'library', 'weld-library')
    dataframe, _, _ = _source_dataframe(source, {'*': WeldLibraryRow})
    if dataframe.empty:
        return False
    changed = False
    if update_segment_count:
        segment_count = _prefab_segment_count(dataframe)
        next_value = str(segment_count)
        if project.pipe_segment != next_value:
            project.pipe_segment = next_value
            changed = True

    if update_prefab_weld_count:
        next_count = int(len(dataframe))
        if project.prefab_weld_count != next_count:
            project.prefab_weld_count = next_count
            changed = True

    if update_completion_rate:
        completed_col = _resolve_column(dataframe.columns, WELD_COLUMNS['completed_flag'])
        total = len(dataframe)
        completed_count = int(dataframe[completed_col].map(_truthy_text).sum()) if completed_col and total else 0
        next_rate = Decimal('0.00') if total == 0 else Decimal(str(round(completed_count / total * 100, 2))).quantize(Decimal('0.01'))
        if project.completion_rate != next_rate:
            project.completion_rate = next_rate
            changed = True

    if changed:
        project.save(update_fields=['pipe_segment', 'prefab_weld_count', 'completion_rate', 'updated_at'])
    return changed


def _sync_completed_plan_rows_to_weld_library(project, rows):
    if not rows:
        return 0

    source = _latest_source(project, 'library', 'weld-library')
    library_df, selected_sheet, _ = _source_dataframe(source, {'*': WeldLibraryRow})
    if library_df.empty:
        return 0

    plan_df = pd.DataFrame(rows)
    plan_completed_col = _resolve_column(plan_df.columns, WELD_COLUMNS['completed_flag'])
    if not plan_completed_col:
        return 0

    library_completed_col = _resolve_column(library_df.columns, WELD_COLUMNS['completed_flag'])
    if not library_completed_col:
        library_completed_col = WELD_COLUMNS['completed_flag']
        library_df[library_completed_col] = False

    plan_df = plan_df.copy()
    plan_df['_key'] = _weld_key_series(plan_df)
    completion_map = (
        plan_df.loc[plan_df['_key'].astype(str).str.len() > 0]
        .drop_duplicates('_key', keep='last')
        .set_index('_key')[plan_completed_col]
        .map(_truthy_text)
        .to_dict()
    )
    if not completion_map:
        return 0

    library_df = library_df.copy()
    library_df['_key'] = _weld_key_series(library_df)
    matched_mask = library_df['_key'].isin(completion_map)
    changed_count = int(matched_mask.sum())
    if changed_count <= 0:
        return 0

    library_df.loc[matched_mask, library_completed_col] = library_df.loc[matched_mask, '_key'].map(completion_map)
    library_df = library_df.drop(columns=['_key'])
    replace_source_rows(
        project,
        'library',
        'weld-library',
        source.display_name if source else WELD_LIBRARY_FILE_NAME,
        source.relative_path if source else 'database://library/weld-library/预制焊口库.xlsx',
        {selected_sheet or 'Sheet1': {'columns': list(library_df.columns), 'rows': library_df.to_dict(orient='records')}},
        {'*': WeldLibraryRow},
    )
    return changed_count


def _sync_anti_corrosion_completed_rows_to_weld_library(project, rows):
    if not rows:
        return 0

    source = _latest_source(project, 'library', 'weld-library')
    library_df, selected_sheet, _ = _source_dataframe(source, {'*': WeldLibraryRow})
    if library_df.empty:
        return 0

    plan_df = pd.DataFrame(rows)
    plan_completed_col = _resolve_column(plan_df.columns, WELD_COLUMNS['material_anti_corrosion_status'])
    if not plan_completed_col:
        return 0

    library_completed_col = _resolve_column(library_df.columns, WELD_COLUMNS['material_anti_corrosion_status'])
    if not library_completed_col:
        library_completed_col = WELD_COLUMNS['material_anti_corrosion_status']
        library_df[library_completed_col] = False

    plan_df = plan_df.copy()
    plan_df['_key'] = _weld_key_series(plan_df)
    completion_map = (
        plan_df.loc[plan_df['_key'].astype(str).str.len() > 0]
        .drop_duplicates('_key', keep='last')
        .set_index('_key')[plan_completed_col]
        .map(_truthy_text)
        .to_dict()
    )
    if not completion_map:
        return 0

    library_df = library_df.copy()
    library_df['_key'] = _weld_key_series(library_df)
    matched_mask = library_df['_key'].isin(completion_map)
    changed_count = int(matched_mask.sum())
    if changed_count <= 0:
        return 0

    library_df.loc[matched_mask, library_completed_col] = library_df.loc[matched_mask, '_key'].map(completion_map)
    library_df = library_df.drop(columns=['_key'])
    replace_source_rows(
        project,
        'library',
        'weld-library',
        source.display_name if source else WELD_LIBRARY_FILE_NAME,
        source.relative_path if source else 'database://library/weld-library/预制焊口库.xlsx',
        {selected_sheet or 'Sheet1': {'columns': list(library_df.columns), 'rows': library_df.to_dict(orient='records')}},
        {'*': WeldLibraryRow},
    )
    return changed_count


def _stage_root(project, token):
    safe_token = str(token or '').strip()
    if not re.fullmatch(r'[0-9a-fA-F-]{32,36}', safe_token):
        raise ValueError('暂存令牌无效')
    return STAGED_PLAN_ROOT / _safe_project_dir_name(project) / safe_token


def _arrival_file_payload(file_path):
    stat = file_path.stat()
    return {
        'name': file_path.name,
        'path': _relative_path(file_path),
        'size': stat.st_size,
        'updatedAt': stat.st_mtime,
    }


def _arrival_rows_summary(rows):
    actual_qty_col = '实际到货数量'
    send_qty_col = '发货数量（米/根）'
    pipe_count_col = '实际到货支数'
    category_col = '材料分类'
    unit_col = '单位'

    total_qty = 0
    pipe_rows = 0
    fitting_rows = 0
    pipe_count = 0
    categories = {}

    for row in rows or []:
        qty = pd.to_numeric(row.get(actual_qty_col), errors='coerce')
        if pd.isna(qty) or float(qty) <= 0:
            qty = pd.to_numeric(row.get(send_qty_col), errors='coerce')
        if pd.notna(qty):
            total_qty += float(qty)

        category = str(row.get(category_col, '') or '').strip() or '未分类'
        unit = str(row.get(unit_col, '') or '').strip()
        name = str(row.get('名称', '') or '').strip()
        is_pipe = unit == '米' or name == '管子' or category == '直管'
        if is_pipe:
            pipe_rows += 1
            count = pd.to_numeric(row.get(pipe_count_col), errors='coerce')
            if pd.notna(count):
                pipe_count += int(count)
        else:
            fitting_rows += 1

        categories[category] = categories.get(category, 0) + 1

    return {
        'materialRows': len(rows or []),
        'pipeRows': pipe_rows,
        'fittingRows': fitting_rows,
        'pipeCount': pipe_count,
        'totalQuantity': round(total_qty, 3),
        'categories': [
            {'category': category, 'count': count}
            for category, count in sorted(categories.items(), key=lambda item: item[0])
        ],
    }


def _parse_library_save_payload(request):
    try:
        payload = json.loads(request.body.decode('utf-8'))
    except (UnicodeDecodeError, json.JSONDecodeError):
        return None, '请求内容不是有效 JSON'

    columns = payload.get('columns')
    rows = payload.get('rows')
    sheet = payload.get('sheet') or ''
    if not isinstance(columns, list) or not columns or not all(isinstance(column, str) for column in columns):
        return None, '列定义无效'
    if not isinstance(rows, list) or not all(isinstance(row, dict) for row in rows):
        return None, '表格数据无效'
    if not isinstance(sheet, str):
        return None, '工作表名称无效'

    return {
        'sheet': sheet.strip(),
        'columns': columns,
        'rows': rows,
    }, None


def _format_schedule_date(value):
    text = str(value or '').strip()
    if re.fullmatch(r'\d{8}', text):
        return f'{text[:4]}-{text[4:6]}-{text[6:]}'
    return text


def _next_master_welding_date(project):
    if project is None:
        return ''
    try:
        ensure_project_tables(project)
        with using_project_tables(project):
            row = (
                MasterScheduleRow.objects
                .filter(project=project)
                .exclude(weld_date='')
                .filter(weld_order_no='')
                .order_by('weld_date', 'source_sheet', 'library_seq')
                .first()
            )
            return _format_schedule_date(row.weld_date) if row else ''
    except Exception:
        return ''


def _next_master_cutting_date(project):
    if project is None:
        return ''
    try:
        ensure_project_tables(project)
        with using_project_tables(project):
            row = (
                MasterScheduleRow.objects
                .filter(project=project)
                .exclude(cut_date='')
                .filter(cut_order_no='')
                .order_by('cut_date', 'source_sheet', 'library_seq')
                .first()
            )
            return _format_schedule_date(row.cut_date) if row else ''
    except Exception:
        return ''


def _schedule_policy_defaults(project):
    if project is None:
        return {}
    try:
        ensure_project_tables(project)
        with using_project_tables(project):
            policy = ProjectSchedulePolicy.objects.filter(project=project).first()
            if not policy:
                return {}
            return {
                'targetDiameter': float(policy.target_diameter),
                'ordersPerDay': int(policy.orders_per_day),
                'skipHolidays': bool(policy.skip_holidays),
                'holidayDates': ', '.join(policy.holiday_dates or []),
                'canceledWeekendDates': ', '.join(policy.canceled_weekend_dates or []),
                'cuttingLeadDays': int(policy.cutting_lead_days),
                'antiCorrosionLeadDays': int(policy.anti_corrosion_lead_days),
            }
    except Exception:
        return {}


def _welding_schedule_defaults(project=None):
    defaults = {
        'weldDate': _next_master_welding_date(project) or datetime.now().strftime('%Y-%m-%d'),
        'weldStartDate': _next_master_welding_date(project) or datetime.now().strftime('%Y-%m-%d'),
        'dateMode': 'auto',
        'manualWeldDates': '',
        'maxDays': '',
        'skipHolidays': True,
        'holidayDates': '',
        'canceledWeekendDates': '',
        'targetDiameter': 260,
        'ordersPerDay': 3,
    }
    try:
        auto_weld_dir = PREFAB_ROOT / 'welding' / 'auto_weld_schedule'
        for path in (PREFAB_ROOT, auto_weld_dir):
            if str(path) not in sys.path:
                sys.path.insert(0, str(path))
        from welding.weld_config import EXTRACT
        defaults['targetDiameter'] = EXTRACT.get('target_diameter', defaults['targetDiameter'])
        defaults['ordersPerDay'] = EXTRACT.get('num_extractions', defaults['ordersPerDay'])
    except Exception:
        pass
    defaults.update({
        key: value
        for key, value in _schedule_policy_defaults(project).items()
        if value not in (None, '')
    })
    return defaults


def _cutting_schedule_defaults(project=None):
    defaults = {
        **_welding_schedule_defaults(project),
        'weldStartDate': _next_master_cutting_date(project) or _welding_schedule_defaults(project).get('weldStartDate') or datetime.now().strftime('%Y-%m-%d'),
    }
    return defaults


def _future_schedule_defaults(project=None):
    defaults = {
        **_welding_schedule_defaults(project),
        'weldStartDate': (datetime.now().date() + timedelta(days=1)).strftime('%Y-%m-%d'),
        'maxDays': '',
        'dateMode': 'auto',
        'manualWeldDates': '',
        'skipHolidays': True,
        'holidayDates': '',
        'canceledWeekendDates': '',
        'cuttingLeadDays': 1,
        'antiCorrosionLeadDays': 1,
        'commissionArea': 1500,
    }
    defaults.update(_schedule_policy_defaults(project))
    return defaults


def _action_payload(action_key, project=None):
    action = ACTIONS[action_key]
    payload = {**action, 'key': action_key, 'script': _relative_path(action['script'])}
    if action_key == 'auto-weld-schedule':
        payload['defaults'] = _welding_schedule_defaults(project)
    if action_key == 'cutting-schedule':
        payload['defaults'] = _cutting_schedule_defaults(project)
    if action_key == 'future-schedule':
        payload['defaults'] = _future_schedule_defaults(project)
    if action_key == 'prefab-weld-library':
        payload['initializationFilters'] = [
            {'key': 'prefabWeldArea', 'stage': '可预制焊口过滤', 'field': '焊口区域', 'operator': '等于', 'value': 'S', 'defaultEnabled': True},
            {'key': 'prefabMaterialType', 'stage': '可预制焊口过滤', 'field': '材料类型', 'operator': '等于', 'value': 'CS', 'defaultEnabled': True},
            {'key': 'autoJointType', 'stage': '自动焊口过滤', 'field': '连接类型', 'operator': '等于', 'value': 'BW', 'defaultEnabled': True},
            {'key': 'autoWallThickness', 'stage': '自动焊口过滤', 'field': '壁厚', 'operator': '介于（含边界）', 'value': '6 ～ 25', 'defaultEnabled': True},
            {'key': 'autoDiameter', 'stage': '自动焊口过滤', 'field': '寸径', 'operator': '介于（含边界）', 'value': '8 ～ 24', 'defaultEnabled': True},
            {'key': 'autoSegmentNo', 'stage': '自动焊口过滤', 'field': '管段号', 'operator': '排除', 'value': '空值', 'defaultEnabled': True},
        ]
    return payload


def _module_payload(module, data_root=DATA_ROOT, project=None):
    file_infos = [_database_module_file_info(project, file_name, data_root) for file_name in module['files']]
    return {
        **module,
        'actions': [_action_payload(action_key, project) for action_key in module['actions']],
        'files': file_infos,
        'readyCount': sum(1 for item in file_infos if item['exists']),
        'totalCount': len(file_infos),
    }


def _decode_process_output(output):
    if not output:
        return ''

    encodings = [
        'utf-8',
        'utf-8-sig',
        locale.getpreferredencoding(False),
        'gb18030',
        'gbk',
    ]
    for encoding in dict.fromkeys(item for item in encodings if item):
        try:
            return output.decode(encoding)
        except UnicodeDecodeError:
            continue
    return output.decode('utf-8', errors='replace')




__all__ = [name for name in globals() if not name.startswith('__')]

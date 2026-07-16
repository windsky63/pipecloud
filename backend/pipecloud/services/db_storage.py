import json
import time
from decimal import Decimal, InvalidOperation
from pathlib import Path

from django.db import models
from django.db import transaction
from django.db.models import Q
from openpyxl import load_workbook

from pipecloud.models import (
    AntiCorrosionMaterialOrderRow,
    ArrivalMaterialRow,
    ArrivalOrderRow,
    DataSourceFile,
    FittingMaterialRow,
    InitializationMaterialRow,
    InitializationWeldRow,
    InitializationWeldExtraData,
    MasterScheduleRow,
    MaterialMatchDetailRow,
    PipeMaterialRow,
    WeldingPlanRow,
    WeldCommonData,
    WeldLibraryRow,
    WeldPreScheduleRow,
    WeldStatusRow,
)
from pipecloud.services.project_tables import ensure_project_tables, using_project_tables


SYSTEM_FIELDS = {
    'id',
    'project',
    'source_file',
    'common_data',
    'sheet_name',
    'row_index',
    'created_at',
    'updated_at',
}

COLUMN_ALIASES = {
    '单元号(必填)': 'unit',
    '管线号(必填)': 'pipeline',
    '预制组件': 'segment_no',
    '预制管段': 'segment_no',
    '预制段': 'segment_no',
    '焊口号': 'weld_no_final',
    '焊接类型': 'joint_type',
    '连接方式': 'joint_type',
    '壁厚尺寸': 'wall_thickness',
    '英制': 'diameter',
    '英制尺寸': 'diameter',
    '公制外径': 'outer_diameter',
    '焊接区': 'weld_area',
    '材料编号1': 'material_mark_1',
    '材料编号2': 'material_mark_2',
    '数量_1': 'quantity_1',
    '数量_2': 'quantity_2',
    '材料唯一码_1': 'material_unique_1',
    '材料唯一码_2': 'material_unique_2',
    '材料唯一编码1': 'material_unique_1',
    '材料唯一编码2': 'material_unique_2',
    '材质名称': 'material',
    '材质类型': 'material_type',
    '油漆1': 'material_paint_1',
    '油漆2': 'material_paint_2',
    '材料描述1': 'description_1',
    '材料描述2': 'description_2',
    '到货状态': 'material_arrival_status',
    '防腐状态': 'material_anti_corrosion_status',
    '下料状态': 'material_cutting_status',
    '材料焊接状态': 'completed_flag',
    '焊接状态': 'completed_flag',
    '是否完成': 'completed_flag',
    '完成状态': 'completed_flag',
    '完工状态': 'completed_flag',
    '材料代码（NCC文本）': 'material_code_ncc',
    '材料代码': 'material_code',
    '库存数量（米）': 'stock_qty',
    '发货数量（米/根）': 'shipment_qty',
}

TRUE_VALUES = {'1', 'true', 't', 'yes', 'y', '是', '已完成', '完成'}
FALSE_VALUES = {'0', 'false', 'f', 'no', 'n', '否', '未完成', '未', ''}
WELD_STATUS_FIELD_LABELS = {
    'priority': '优先级',
    'material_arrival_status': '材料到货状态',
    'material_anti_corrosion_status': '材料防腐状态',
    'material_cutting_status': '材料下料状态',
    'completed_flag': '材料焊接状态',
}
WELD_STATUS_COMPLETION_ALIASES = {'是否完成', '完成状态', '完工状态', '焊接状态'}
WELD_STATUS_SOURCE_MODELS = {
    WeldLibraryRow,
    WeldPreScheduleRow,
    MasterScheduleRow,
    WeldingPlanRow,
}
WELD_COMMON_DATA_SOURCE_MODELS = {
    InitializationWeldRow,
    WeldLibraryRow,
    WeldPreScheduleRow,
    MasterScheduleRow,
    WeldingPlanRow,
}
WELD_COMMON_DATA_FIELDS = {
    'wall_thickness',
    'wall_thickness_no',
    'diameter',
    'outer_diameter',
    'material',
    'material_code_name',
    'material_type',
    'material_mark_1',
    'material_mark_2',
    'material_unique_1',
    'material_unique_2',
    'material_code_1',
    'material_code_2',
    'material_paint_1',
    'material_paint_2',
    'quantity_1',
    'quantity_2',
    'material_unit_1',
    'material_unit_2',
    'unit_name_1',
    'unit_name_2',
    'description_1',
    'description_2',
}
WELD_COMMON_DATA_FIELD_ALIASES = {
    'material_type': 'material_code_name',
    'unit_name_1': 'material_unit_1',
    'unit_name_2': 'material_unit_2',
}
WELD_COMMON_DATA_MODEL_FIELDS = {
    field.name
    for field in WeldCommonData._meta.fields
}


def model_column_map(model):
    mapping = {}
    model_field_names = {
        field.name
        for field in model._meta.fields
        if field.name not in SYSTEM_FIELDS
    }
    for field in model._meta.fields:
        if field.name in SYSTEM_FIELDS:
            continue
        label = str(field.verbose_name)
        mapping[label] = field.name
    mapping.update({key: value for key, value in COLUMN_ALIASES.items() if value in model_field_names})
    if 'weld_method' in model_field_names:
        mapping['焊接方法'] = 'weld_method'
    if 'welding_mode' in model_field_names:
        mapping['焊接方式'] = 'welding_mode'
    return mapping


def model_columns(model):
    return [
        str(field.verbose_name)
        for field in model._meta.fields
        if field.name not in SYSTEM_FIELDS
    ]


def model_field_labels(model):
    return {
        field.name: str(field.verbose_name)
        for field in model._meta.fields
        if field.name not in SYSTEM_FIELDS
    }


def project_uses_material_units(project):
    if WeldCommonData.objects.filter(project=project).filter(
        Q(material_unit_1__gt='') | Q(material_unit_2__gt='')
    ).exists():
        return True
    return InitializationWeldRow.objects.filter(project=project).filter(
        Q(material_unit_1__gt='') | Q(material_unit_2__gt='')
    ).exists()


def clean_cell(value):
    if value is None:
        return ''
    if isinstance(value, Decimal):
        normalized = value.normalize()
        if normalized == normalized.to_integral():
            return str(normalized.quantize(Decimal('1')))
        return format(normalized, 'f').rstrip('0').rstrip('.')
    if hasattr(value, 'isoformat'):
        return value.isoformat()
    return str(value)


def _clean_status_value(value):
    text = clean_cell(value).strip()
    return '' if text.lower() == 'nan' else text


def _weld_status_values_for_project(project, library_seqs):
    seqs = {
        str(value or '').strip()
        for value in (library_seqs or [])
        if str(value or '').strip()
    }
    if not seqs:
        return {}
    return {
        row.library_seq: {
            '优先级': row.priority,
            '材料到货状态': row.material_arrival_status,
            '材料防腐状态': row.material_anti_corrosion_status,
            '材料下料状态': row.material_cutting_status,
            '材料焊接状态': row.completed_flag,
            '是否完成': row.completed_flag,
            '完成状态': row.completed_flag,
            '完工状态': row.completed_flag,
            '焊接状态': row.completed_flag,
        }
        for row in WeldStatusRow.objects.filter(project=project, library_seq__in=seqs)
    }


def apply_weld_status_to_rows(project, columns, rows):
    if not rows or '库序号' not in columns:
        return rows
    status_by_seq = _weld_status_values_for_project(project, [row.get('库序号') for row in rows])
    if not status_by_seq:
        return rows
    status_columns = set(WELD_STATUS_FIELD_LABELS.values()) | WELD_STATUS_COMPLETION_ALIASES
    next_rows = []
    for row in rows:
        next_row = dict(row)
        status = status_by_seq.get(str(row.get('库序号') or '').strip())
        if status:
            for column in columns:
                if column in status_columns and status.get(column) not in (None, ''):
                    next_row[column] = status[column]
        next_rows.append(next_row)
    return next_rows


def apply_weld_status_to_dataframe(project, dataframe):
    if dataframe is None or dataframe.empty or '库序号' not in dataframe.columns:
        return dataframe
    rows = apply_weld_status_to_rows(project, list(dataframe.columns), dataframe.fillna('').to_dict('records'))
    return dataframe.__class__(rows, columns=list(dataframe.columns))


def _sync_weld_status_values(project, status_rows):
    updates = {}
    for item in status_rows:
        library_seq = str(item.get('library_seq') or '').strip()
        if not library_seq:
            continue
        target = updates.setdefault(library_seq, {})
        for field_name in WELD_STATUS_FIELD_LABELS:
            if field_name in item:
                value = item.get(field_name)
                if field_name == 'priority':
                    target[field_name] = _clean_status_value(value)
                    continue
                if value in (None, ''):
                    continue
                target[field_name] = coerce_boolean(value)
    updates = {library_seq: values for library_seq, values in updates.items() if values}
    if not updates:
        return 0

    existing = {
        row.library_seq: row
        for row in WeldStatusRow.objects.filter(project=project, library_seq__in=updates.keys())
    }
    missing = [
        WeldStatusRow(project=project, library_seq=library_seq, **values)
        for library_seq, values in updates.items()
        if library_seq not in existing
    ]
    if missing:
        WeldStatusRow.objects.bulk_create(missing, batch_size=500, ignore_conflicts=True)
        existing = {
            row.library_seq: row
            for row in WeldStatusRow.objects.filter(project=project, library_seq__in=updates.keys())
        }
    changed_count = len(missing)

    changed_rows = []
    update_fields = set()
    for library_seq, values in updates.items():
        row = existing.get(library_seq)
        if row is None:
            continue
        changed = []
        for field_name, value in values.items():
            if getattr(row, field_name) != value:
                setattr(row, field_name, value)
                update_fields.add(field_name)
                changed.append(field_name)
        if changed:
            changed_rows.append(row)
    if changed_rows and update_fields:
        WeldStatusRow.objects.bulk_update(changed_rows, sorted(update_fields), batch_size=500)
        changed_count += len(changed_rows)
    return changed_count


def _sync_weld_common_data(project, values):
    library_seq, defaults = _weld_common_defaults(values)
    if not library_seq:
        return None
    common_data, created = WeldCommonData.objects.get_or_create(
        project=project,
        library_seq=library_seq,
        defaults=defaults,
    )
    if created:
        return common_data
    changed = []
    for field_name, value in defaults.items():
        if getattr(common_data, field_name) != value:
            setattr(common_data, field_name, value)
            changed.append(field_name)
    if changed:
        common_data.save(update_fields=[*changed, 'updated_at'])
    return common_data


def _weld_common_defaults(values):
    library_seq = _clean_status_value(values.get('library_seq'))
    if not library_seq:
        return '', {}
    defaults = {}
    for source_field in WELD_COMMON_DATA_FIELDS:
        target_field = WELD_COMMON_DATA_FIELD_ALIASES.get(source_field, source_field)
        if target_field not in WELD_COMMON_DATA_MODEL_FIELDS:
            continue
        value = _clean_status_value(values.get(source_field))
        if value:
            defaults[target_field] = value
    return library_seq, defaults


def _sync_weld_common_data_many(project, values_list, update_existing=True):
    defaults_by_seq = {}
    for values in values_list:
        library_seq, defaults = _weld_common_defaults(values)
        if not library_seq:
            continue
        target = defaults_by_seq.setdefault(library_seq, {})
        target.update(defaults)
    if not defaults_by_seq:
        return {}

    existing = {
        item.library_seq: item
        for item in WeldCommonData.objects.filter(project=project, library_seq__in=defaults_by_seq.keys())
    }
    missing = [
        WeldCommonData(project=project, library_seq=library_seq, **defaults)
        for library_seq, defaults in defaults_by_seq.items()
        if library_seq not in existing
    ]
    if missing:
        WeldCommonData.objects.bulk_create(missing, batch_size=500, ignore_conflicts=True)
        existing = {
            item.library_seq: item
            for item in WeldCommonData.objects.filter(project=project, library_seq__in=defaults_by_seq.keys())
        }

    if not update_existing:
        return existing

    changed_items = []
    update_fields = set()
    for library_seq, defaults in defaults_by_seq.items():
        common_data = existing.get(library_seq)
        if common_data is None:
            continue
        changed = False
        for field_name, value in defaults.items():
            if getattr(common_data, field_name) != value:
                setattr(common_data, field_name, value)
                update_fields.add(field_name)
                changed = True
        if changed:
            changed_items.append(common_data)
    if changed_items and update_fields:
        WeldCommonData.objects.bulk_update(changed_items, sorted(update_fields), batch_size=500)
    return existing


def coerce_boolean(value):
    if isinstance(value, bool):
        return value
    if value is None:
        return False
    if isinstance(value, (int, float)):
        return value != 0
    text = str(value).strip().lower()
    if text in TRUE_VALUES:
        return True
    if text in FALSE_VALUES:
        return False
    return bool(text)


def coerce_model_value(field, value):
    if isinstance(field, models.BooleanField):
        return coerce_boolean(value)
    if isinstance(field, models.DecimalField):
        if value is None or value == '':
            return Decimal('0')
        try:
            number = Decimal(str(value).strip())
        except (InvalidOperation, ValueError):
            return Decimal('0')
        quant = Decimal('1').scaleb(-field.decimal_places)
        return number.quantize(quant)
    return clean_cell(value)


def normalize_json_value(value, fallback):
    if value is None or value == '':
        return fallback
    if isinstance(value, str):
        try:
            parsed = json.loads(value)
        except json.JSONDecodeError:
            return fallback
        return parsed
    return value


def row_values(columns, values):
    padded = list(values or [])[:len(columns)]
    if len(padded) < len(columns):
        padded.extend([''] * (len(columns) - len(padded)))
    return {column: clean_cell(value) for column, value in zip(columns, padded)}


def relative_path(path, backend_dir):
    try:
        return str(Path(path).resolve().relative_to(Path(backend_dir).resolve())).replace('\\', '/')
    except ValueError:
        return str(path).replace('\\', '/')


def read_workbook_rows(file_path):
    workbook = load_workbook(file_path, read_only=True, data_only=True)
    try:
        payload = {}
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            rows_iter = worksheet.iter_rows(values_only=True)
            header = next(rows_iter, [])
            columns = [str(column or '') for column in header]
            rows = [row_values(columns, values) for values in rows_iter]
            payload[sheet_name] = {'columns': columns, 'rows': rows}
        return payload
    finally:
        workbook.close()


INITIALIZATION_REQUIRED_FIELDS = {
    'unit',
    'pipeline',
    'diameter',
    'joint_type',
}
INITIALIZATION_REQUIRED_FIELD_GROUPS = (
    ('weld_no_start', 'weld_no_final'),
)

INITIALIZATION_WELD_SHEET = '焊口表'
INITIALIZATION_MATERIAL_SHEET = '材料表'
INITIALIZATION_MATERIAL_MODELS = {'*': InitializationMaterialRow}

INITIALIZATION_COMPATIBILITY_COLUMNS = {
    '焊口号': '最终焊口号',
    '焊接类型': '接头类型',
    '材料描述1': '描述1',
    '材料描述2': '描述2',
}


def _source_column_mapping(model, columns):
    column_map = model_column_map(model)
    mapping = {}
    aliases = []
    for column in columns:
        source_label = str(column or '').strip()
        field_name = column_map.get(source_label)
        if not field_name or field_name in mapping:
            continue
        target_label = model_field_labels(model).get(field_name, source_label)
        mapping[field_name] = source_label
        if source_label != target_label:
            aliases.append({
                'source': source_label,
                'target': target_label,
                'field': field_name,
            })
    return mapping, aliases


def normalize_rows_for_model(model, columns, rows):
    field_labels = model_field_labels(model)
    source_mapping, aliases = _source_column_mapping(model, columns)
    normalized_columns = list(field_labels.values())
    normalized_rows = []
    for row in rows:
        normalized_row = {}
        for field_name, label in field_labels.items():
            source_column = source_mapping.get(field_name)
            normalized_row[label] = row.get(source_column, '') if source_column else ''
        normalized_rows.append(normalized_row)
    missing_fields = [
        {'field': field_name, 'column': label}
        for field_name, label in field_labels.items()
        if field_name not in source_mapping
    ]
    return normalized_columns, normalized_rows, aliases, missing_fields, source_mapping


def initialization_rows_with_compatibility(rows):
    compatible_rows = []
    for row in rows:
        compatible = dict(row)
        for target, source in INITIALIZATION_COMPATIBILITY_COLUMNS.items():
            if target not in compatible or compatible.get(target, '') == '':
                compatible[target] = compatible.get(source, '')
        compatible_rows.append(compatible)
    return compatible_rows


def normalize_initialization_payload(project, workbook_payload, include_visible_rows=True):
    weld_sheet = INITIALIZATION_WELD_SHEET if INITIALIZATION_WELD_SHEET in workbook_payload else ''
    if not weld_sheet:
        weld_sheet = next(
            (
                name for name in workbook_payload
                if name not in {INITIALIZATION_MATERIAL_SHEET, '解析概况'}
            ),
            next(iter(workbook_payload), 'Sheet1'),
        )
    source_payload = workbook_payload.get(weld_sheet) or {'columns': [], 'rows': []}
    source_columns = [str(column or '').strip() for column in source_payload.get('columns') or []]
    source_rows = source_payload.get('rows') or []
    core_columns, core_rows, aliases, missing_fields, source_mapping = normalize_rows_for_model(
        InitializationWeldRow,
        source_columns,
        source_rows,
    )

    missing_required, invalid_rows = validate_normalized_rows(
        InitializationWeldRow,
        core_rows,
        source_mapping,
        INITIALIZATION_REQUIRED_FIELDS,
    )
    mapped_source_columns = set(source_mapping.values())
    extra_columns = [
        column for column in source_columns
        if column and column not in mapped_source_columns
    ]

    prefix = f'P{project.id}-W'
    next_counter = 0
    existing_sequences = InitializationWeldRow.objects.filter(
        project=project,
    ).values_list('library_seq', flat=True)
    for value in existing_sequences:
        text = str(value or '')
        if not text.startswith(prefix):
            continue
        try:
            next_counter = max(next_counter, int(text[len(prefix):]))
        except ValueError:
            continue
    reserved = set()
    normalized_rows = []
    extra_rows = []
    for source_row, core_row in zip(source_rows, core_rows):
        requested = str(core_row.get('库序号', '') or '').strip()
        sequence = requested if requested and requested not in reserved else ''
        if not sequence:
            while True:
                next_counter += 1
                sequence = f'{prefix}{next_counter:08d}'
                if sequence not in reserved:
                    break
        reserved.add(sequence)
        core_row['库序号'] = sequence
        normalized_rows.append(core_row)
        extra_rows.append({
            column: clean_cell(source_row.get(column, ''))
            for column in extra_columns
            if clean_cell(source_row.get(column, '')) != ''
        })

    columns = [*core_columns, *extra_columns]
    visible_rows = []
    if include_visible_rows:
        for core_row, extra_row in zip(normalized_rows, extra_rows):
            visible_rows.append({**core_row, **extra_row})
    validation = {
        'standardized': True,
        'canImport': not missing_required and not invalid_rows,
        'sheets': [{
            'sheet': weld_sheet,
            'rowCount': len(normalized_rows),
            'sourceColumnCount': len(source_columns),
            'standardColumnCount': len(core_columns),
            'extraColumnCount': len(extra_columns),
            'aliasMappings': aliases,
            'missingFields': missing_fields,
            'missingRequiredFields': missing_required,
            'invalidRows': invalid_rows,
            'canImport': not missing_required and not invalid_rows,
        }],
    }
    return {
        'sheet': weld_sheet,
        'columns': columns,
        'fixedColumns': core_columns,
        'extraColumns': extra_columns,
        'coreRows': normalized_rows,
        'extraRows': extra_rows,
        'visibleRows': visible_rows,
        'validation': validation,
    }


def validate_normalized_rows(model, rows, source_mapping, required_fields):
    field_labels = model_field_labels(model)
    required = [field_name for field_name in required_fields if field_name in field_labels]
    missing_required = [
        {'field': field_name, 'column': field_labels[field_name]}
        for field_name in required
        if field_name not in source_mapping
    ]
    required_groups = INITIALIZATION_REQUIRED_FIELD_GROUPS if model is InitializationWeldRow else ()
    for group in required_groups:
        group_fields = [field_name for field_name in group if field_name in field_labels]
        if group_fields and not any(field_name in source_mapping for field_name in group_fields):
            missing_required.append({
                'field': '|'.join(group_fields),
                'column': ' / '.join(field_labels[field_name] for field_name in group_fields),
            })
    invalid_rows = []
    for index, row in enumerate(rows, start=1):
        empty_columns = [
            field_labels[field_name]
            for field_name in required
            if not str(row.get(field_labels[field_name], '') or '').strip()
        ]
        for group in required_groups:
            group_fields = [field_name for field_name in group if field_name in field_labels]
            if group_fields and not any(
                str(row.get(field_labels[field_name], '') or '').strip()
                for field_name in group_fields
            ):
                empty_columns.append(' / '.join(field_labels[field_name] for field_name in group_fields))
        if empty_columns:
            invalid_rows.append({
                'rowIndex': index,
                'missingColumns': empty_columns,
            })
        if len(invalid_rows) >= 20:
            break
    return missing_required, invalid_rows


def standardize_workbook_payload(workbook_payload, sheet_models, required_fields=None):
    required_fields = set(required_fields or [])
    normalized_payload = {}
    sheet_summaries = []
    can_import = True
    for sheet_name, payload in workbook_payload.items():
        model = sheet_models.get(sheet_name) or sheet_models.get('*')
        if model is None:
            normalized_payload[sheet_name] = payload
            continue

        columns, rows, aliases, missing_fields, source_mapping = normalize_rows_for_model(
            model,
            payload.get('columns') or [],
            payload.get('rows') or [],
        )
        missing_required, invalid_rows = validate_normalized_rows(model, rows, source_mapping, required_fields)
        if missing_required:
            can_import = False
        normalized_payload[sheet_name] = {'columns': columns, 'rows': rows}
        sheet_summaries.append({
            'sheet': sheet_name,
            'rowCount': len(rows),
            'sourceColumnCount': len(payload.get('columns') or []),
            'standardColumnCount': len(columns),
            'aliasMappings': aliases,
            'missingFields': missing_fields,
            'missingRequiredFields': missing_required,
            'invalidRows': invalid_rows,
            'canImport': not missing_required and not invalid_rows,
        })
    return normalized_payload, {
        'standardized': True,
        'canImport': can_import,
        'sheets': sheet_summaries,
    }


def initialization_preview_payload(file_path, sheet_name=None, limit=20, project=None):
    workbook_payload = read_workbook_rows(file_path)
    if project is not None:
        normalized = normalize_initialization_payload(project, workbook_payload)
        rows = normalized['visibleRows']
        return {
            'sheet': normalized['sheet'],
            'sheets': [normalized['sheet']],
            'total': len(rows),
            'columns': normalized['columns'],
            'rows': rows[:limit],
            'fixedColumns': normalized['fixedColumns'],
            'fixedRows': normalized['coreRows'][:limit],
            'extraColumns': normalized['extraColumns'],
            'extraRows': normalized['extraRows'][:limit],
            'previewLimit': limit,
            'normalization': normalized['validation'],
        }
    normalized_payload, validation = standardize_workbook_payload(
        workbook_payload,
        INITIALIZATION_MODELS,
        INITIALIZATION_REQUIRED_FIELDS,
    )
    sheets = list(normalized_payload.keys())
    selected_sheet = sheet_name if sheet_name in sheets else (sheets[0] if sheets else '')
    selected_payload = normalized_payload.get(selected_sheet) or {'columns': [], 'rows': []}
    rows = selected_payload.get('rows') or []
    return {
        'sheet': selected_sheet,
        'sheets': sheets,
        'total': len(rows),
        'columns': selected_payload.get('columns') or [],
        'rows': rows[:limit],
        'previewLimit': limit,
        'normalization': validation,
    }


def workbook_preview_payload(file_path, sheet_name=None, limit=20):
    workbook_payload = read_workbook_rows(file_path)
    sheets = list(workbook_payload.keys())
    selected_sheet = sheet_name if sheet_name in sheets else (sheets[0] if sheets else '')
    selected_payload = workbook_payload.get(selected_sheet) or {'columns': [], 'rows': []}
    rows = selected_payload.get('rows') or []
    return {
        'sheet': selected_sheet,
        'sheets': sheets,
        'total': len(rows),
        'columns': selected_payload.get('columns') or [],
        'rows': rows[:limit],
        'previewLimit': limit,
        'normalization': {
            'standardized': False,
            'canImport': True,
            'sheets': [],
        },
    }


def source_signature(file_path):
    stat = Path(file_path).stat()
    return int(stat.st_size), float(stat.st_mtime)


def upsert_source_file(project, source_type, source_key, file_path, backend_dir, workbook_payload):
    size, updated_at = source_signature(file_path)
    source, _ = DataSourceFile.objects.update_or_create(
        project=project,
        source_type=source_type,
        source_key=source_key,
        relative_path=relative_path(file_path, backend_dir),
        defaults={
            'display_name': Path(file_path).name,
            'file_size': size,
            'file_updated_at': updated_at,
            'sheet_names': list(workbook_payload.keys()),
            'sheet_columns': {
                sheet_name: sheet_payload['columns']
                for sheet_name, sheet_payload in workbook_payload.items()
            },
        },
    )
    return source


def upsert_virtual_source_file(project, source_type, source_key, display_name, relative_path_value, workbook_payload):
    source, _ = DataSourceFile.objects.update_or_create(
        project=project,
        source_type=source_type,
        source_key=source_key,
        relative_path=str(relative_path_value).replace('\\', '/'),
        defaults={
            'display_name': display_name,
            'file_size': 0,
            'file_updated_at': time.time(),
            'sheet_names': list(workbook_payload.keys()),
            'sheet_columns': {
                sheet_name: sheet_payload['columns']
                for sheet_name, sheet_payload in workbook_payload.items()
            },
        },
    )
    return source


def dataframe_payload(dataframe):
    if dataframe is None:
        columns = []
        rows = []
    else:
        normalized = dataframe.fillna('').copy()
        columns = [str(column) for column in normalized.columns]
        rows = [
            {column: clean_cell(row.get(column, '')) for column in columns}
            for row in normalized.to_dict(orient='records')
        ]
    return {'columns': columns, 'rows': rows}


def normalize_workbook_payload_for_models(workbook_payload, sheet_models):
    normalized_payload = {}
    for sheet_name, payload in workbook_payload.items():
        model = sheet_models.get(sheet_name) or sheet_models.get('*')
        if model is None:
            normalized_payload[sheet_name] = payload
            continue
        columns = payload.get('columns') or []
        rows = payload.get('rows') or []
        normalized_columns, normalized_rows, _, _, _ = normalize_rows_for_model(model, columns, rows)
        normalized_payload[sheet_name] = {
            'columns': normalized_columns,
            'rows': normalized_rows,
        }
    return normalized_payload


def sync_table_payload(
    project,
    source_type,
    source_key,
    display_name,
    relative_path_value,
    workbook_payload,
    sheet_models,
    differential=False,
    update_common_data=True,
    sync_status=True,
):
    ensure_project_tables(project)
    with using_project_tables(project), transaction.atomic():
        workbook_payload = normalize_workbook_payload_for_models(workbook_payload, sheet_models)
        source = upsert_virtual_source_file(
            project,
            source_type,
            source_key,
            display_name,
            relative_path_value,
            workbook_payload,
        )
        if differential:
            active_sheets = set(workbook_payload)
            for model in set(sheet_models.values()):
                if model is not None:
                    model.objects.filter(project=project, source_file=source).exclude(sheet_name__in=active_sheets).delete()
        else:
            for model in set(sheet_models.values()):
                if model is not None:
                    model.objects.filter(project=project, source_file=source).delete()
        for sheet_name, payload in workbook_payload.items():
            model = sheet_models.get(sheet_name) or sheet_models.get('*')
            if model is None:
                continue
            if differential:
                sync_rows_for_model_differential(
                    project,
                    source,
                    model,
                    sheet_name,
                    payload['rows'],
                    update_common_data=update_common_data,
                    sync_status=sync_status,
                )
            else:
                sync_rows_for_model(
                    project,
                    source,
                    model,
                    sheet_name,
                    payload['rows'],
                    clear_existing=False,
                    update_common_data=update_common_data,
                    sync_status=sync_status,
                )
        return source


def sync_dataframes(
    project,
    source_type,
    source_key,
    display_name,
    relative_path_value,
    sheet_dataframes,
    sheet_models,
    differential=False,
    update_common_data=True,
    sync_status=True,
):
    workbook_payload = {
        str(sheet_name): dataframe_payload(dataframe)
        for sheet_name, dataframe in sheet_dataframes.items()
    }
    return sync_table_payload(
        project,
        source_type,
        source_key,
        display_name,
        relative_path_value,
        workbook_payload,
        sheet_models,
        differential=differential,
        update_common_data=update_common_data,
        sync_status=sync_status,
    )


def source_is_current(source, file_path):
    if source is None or not Path(file_path).exists():
        return False
    size, updated_at = source_signature(file_path)
    return source.file_size == size and source.file_updated_at == updated_at


def source_for_file(project, source_type, source_key, file_path, backend_dir):
    return DataSourceFile.objects.filter(
        project=project,
        source_type=source_type,
        source_key=source_key,
        relative_path=relative_path(file_path, backend_dir),
    ).first()


def sync_rows_for_model(
    project,
    source,
    model,
    sheet_name,
    rows,
    clear_existing=True,
    update_common_data=True,
    sync_status=True,
):
    mapping = model_column_map(model)
    supports_common_data = any(field.name == 'common_data' for field in model._meta.fields)
    model_fields = {
        field.name: field
        for field in model._meta.fields
        if field.name not in SYSTEM_FIELDS
    }
    if clear_existing:
        model.objects.filter(project=project, source_file=source, sheet_name=sheet_name).delete()
    batch_size = 500
    for batch_start in range(0, len(rows), batch_size):
        prepared_values = []
        status_updates = []
        for offset, row in enumerate(rows[batch_start:batch_start + batch_size], start=1):
            index = batch_start + offset
            values = {
                field_name: False if isinstance(field, models.BooleanField) else ''
                for field_name, field in model_fields.items()
            }
            for column, field_name in mapping.items():
                if field_name in SYSTEM_FIELDS or column not in row:
                    continue
                field = model_fields.get(field_name)
                if field is None:
                    continue
                value = coerce_model_value(field, row.get(column, ''))
                if isinstance(field, models.BooleanField):
                    values[field_name] = value
                elif value or not values.get(field_name):
                    values[field_name] = value
            if sync_status and model in WELD_STATUS_SOURCE_MODELS:
                status_update = {'library_seq': values.get('library_seq', '')}
                for column, field_name in mapping.items():
                    if field_name in WELD_STATUS_FIELD_LABELS and column in row:
                        raw_value = row.get(column, '')
                        if _clean_status_value(raw_value) != '':
                            status_update[field_name] = raw_value
                if len(status_update) > 1:
                    status_updates.append(status_update)
            prepared_values.append((index, values))

        common_data_by_seq = {}
        if model in WELD_COMMON_DATA_SOURCE_MODELS and supports_common_data:
            common_data_by_seq = _sync_weld_common_data_many(
                project,
                [values for _, values in prepared_values],
                update_existing=update_common_data,
            )

        instances = []
        for index, values in prepared_values:
            if common_data_by_seq:
                values['common_data'] = common_data_by_seq.get(_clean_status_value(values.get('library_seq')))
            instances.append(model(
                project=project,
                source_file=source,
                sheet_name=sheet_name,
                row_index=index,
                **values,
            ))
        if instances:
            model.objects.bulk_create(instances, batch_size=batch_size)
        if sync_status and status_updates:
            _sync_weld_status_values(project, status_updates)


def sync_rows_for_model_differential(
    project,
    source,
    model,
    sheet_name,
    rows,
    update_common_data=True,
    sync_status=True,
):
    mapping = model_column_map(model)
    supports_common_data = any(field.name == 'common_data' for field in model._meta.fields)
    model_fields = {
        field.name: field
        for field in model._meta.fields
        if field.name not in SYSTEM_FIELDS
    }
    prepared_values = []
    status_updates = []
    for index, row in enumerate(rows, start=1):
        values = {
            field_name: False if isinstance(field, models.BooleanField) else ''
            for field_name, field in model_fields.items()
        }
        for column, field_name in mapping.items():
            if field_name in SYSTEM_FIELDS or column not in row:
                continue
            field = model_fields.get(field_name)
            if field is None:
                continue
            value = coerce_model_value(field, row.get(column, ''))
            if isinstance(field, models.BooleanField):
                values[field_name] = value
            elif value or not values.get(field_name):
                values[field_name] = value
        if sync_status and model in WELD_STATUS_SOURCE_MODELS:
            status_update = {'library_seq': values.get('library_seq', '')}
            for column, field_name in mapping.items():
                if field_name in WELD_STATUS_FIELD_LABELS and column in row:
                    raw_value = row.get(column, '')
                    if _clean_status_value(raw_value) != '':
                        status_update[field_name] = raw_value
            if len(status_update) > 1:
                status_updates.append(status_update)
        prepared_values.append((index, values))

    common_data_by_seq = {}
    if model in WELD_COMMON_DATA_SOURCE_MODELS and supports_common_data:
        common_data_by_seq = _sync_weld_common_data_many(
            project,
            [values for _, values in prepared_values],
            update_existing=update_common_data,
        )

    existing_rows = list(
        model.objects
        .filter(project=project, source_file=source, sheet_name=sheet_name)
        .order_by('row_index', 'id')
    )
    existing_by_index = {}
    for existing_row in existing_rows:
        if existing_row.row_index in existing_by_index:
            sync_rows_for_model(
                project,
                source,
                model,
                sheet_name,
                rows,
                clear_existing=True,
                update_common_data=update_common_data,
                sync_status=sync_status,
            )
            return
        existing_by_index[existing_row.row_index] = existing_row

    create_rows = []
    changed_rows = []
    update_fields = set()
    for index, values in prepared_values:
        common_data = None
        if common_data_by_seq:
            common_data = common_data_by_seq.get(_clean_status_value(values.get('library_seq')))
        existing_row = existing_by_index.pop(index, None)
        if existing_row is None:
            create_values = dict(values)
            if supports_common_data:
                create_values['common_data'] = common_data
            create_rows.append(model(
                project=project,
                source_file=source,
                sheet_name=sheet_name,
                row_index=index,
                **create_values,
            ))
            continue

        changed = False
        for field_name, value in values.items():
            if getattr(existing_row, field_name) != value:
                setattr(existing_row, field_name, value)
                update_fields.add(field_name)
                changed = True
        if supports_common_data:
            common_data_id = common_data.pk if common_data is not None else None
            if existing_row.common_data_id != common_data_id:
                existing_row.common_data = common_data
                update_fields.add('common_data')
                changed = True
        if changed:
            changed_rows.append(existing_row)

    if existing_by_index:
        model.objects.filter(pk__in=[row.pk for row in existing_by_index.values()]).delete()
    if create_rows:
        model.objects.bulk_create(create_rows, batch_size=500)
    if changed_rows and update_fields:
        model.objects.bulk_update(changed_rows, sorted(update_fields), batch_size=500)
    if sync_status and status_updates:
        _sync_weld_status_values(project, status_updates)


def sync_excel_table(project, file_path, backend_dir, source_type, source_key, sheet_models, force=False):
    ensure_project_tables(project)
    file_path = Path(file_path)
    with using_project_tables(project):
        source = source_for_file(project, source_type, source_key, file_path, backend_dir)
        if not force and source_is_current(source, file_path):
            return source

        workbook_payload = read_workbook_rows(file_path)
        if source_type == 'initialization':
            workbook_payload, validation = standardize_workbook_payload(
                workbook_payload,
                sheet_models,
                INITIALIZATION_REQUIRED_FIELDS,
            )
            if not validation.get('canImport'):
                raise ValueError('初始化数据标准化校验未通过，存在缺失的关键字段或必填值')
        with transaction.atomic():
            source = upsert_source_file(project, source_type, source_key, file_path, backend_dir, workbook_payload)
            for sheet_name, payload in workbook_payload.items():
                model = sheet_models.get(sheet_name) or sheet_models.get('*')
                if model is None:
                    continue
                sync_rows_for_model(project, source, model, sheet_name, payload['rows'])
        return source


def queryset_rows(source, model, sheet_name, limit=None):
    field_columns = [
        (field.name, str(field.verbose_name))
        for field in model._meta.fields
        if field.name not in SYSTEM_FIELDS
    ]
    sheet_columns = normalize_json_value(source.sheet_columns, {})
    if not isinstance(sheet_columns, dict):
        sheet_columns = {}
    columns = sheet_columns.get(sheet_name) or [column for _, column in field_columns]
    label_to_field = {label: name for name, label in field_columns}
    rows = []
    queryset = model.objects.filter(source_file=source, sheet_name=sheet_name).order_by('row_index')
    if model is InitializationWeldRow:
        queryset = queryset.select_related('extra_data')
    if limit is not None:
        queryset = queryset[:max(int(limit), 0)]
    for item in queryset:
        row = {
            column: getattr(item, label_to_field.get(column, ''), '')
            for column in columns
        }
        if model is InitializationWeldRow:
            try:
                custom_fields = item.extra_data.custom_fields or {}
            except InitializationWeldExtraData.DoesNotExist:
                custom_fields = {}
            for column in columns:
                if column in custom_fields:
                    row[column] = custom_fields[column]
        rows.append(row)
    rows = apply_weld_status_to_rows(source.project, columns, rows)
    return columns, rows


def table_payload(source, sheet_models, sheet_name=None):
    with using_project_tables(source.project_id):
        sheets = normalize_json_value(source.sheet_names, [])
        if not isinstance(sheets, list):
            sheets = []
        selected_sheet = sheet_name if sheet_name in sheets else (sheets[0] if sheets else '')
        model = sheet_models.get(selected_sheet) or sheet_models.get('*')
        if not selected_sheet or model is None:
            return selected_sheet, sheets, 0, [], []
        columns, rows = queryset_rows(source, model, selected_sheet)
        return selected_sheet, sheets, len(rows), columns, rows


def table_preview_payload(source, sheet_models, sheet_name=None, limit=20):
    with using_project_tables(source.project_id):
        sheets = normalize_json_value(source.sheet_names, [])
        if not isinstance(sheets, list):
            sheets = []
        selected_sheet = sheet_name if sheet_name in sheets else (sheets[0] if sheets else '')
        model = sheet_models.get(selected_sheet) or sheet_models.get('*')
        if not selected_sheet or model is None:
            return selected_sheet, sheets, 0, [], []
        total = model.objects.filter(source_file=source, sheet_name=selected_sheet).count()
        columns, rows = queryset_rows(source, model, selected_sheet, limit=limit)
        return selected_sheet, sheets, total, columns, rows


def latest_source(project, source_type, source_key=None, source_key_prefix=None, display_name=None):
    ensure_project_tables(project)
    with using_project_tables(project):
        queryset = DataSourceFile.objects.filter(project=project, source_type=source_type)
        if source_key:
            queryset = queryset.filter(source_key=source_key)
        if source_key_prefix:
            queryset = queryset.filter(source_key__startswith=source_key_prefix)
        if display_name:
            queryset = queryset.filter(display_name=display_name)
        return queryset.order_by('-file_updated_at', '-id').first()


def replace_initialization_source_rows(
    project,
    source_key,
    display_name,
    relative_path_value,
    workbook_payload,
):
    ensure_project_tables(project)
    with using_project_tables(project), transaction.atomic():
        type(project).objects.select_for_update().get(pk=project.pk)
        normalized = normalize_initialization_payload(project, workbook_payload, include_visible_rows=False)
        validation = normalized['validation']
        if not validation.get('canImport'):
            raise ValueError('初始化数据标准化校验未通过，存在缺失的关键字段或必填值')

        material_payload = workbook_payload.get(INITIALIZATION_MATERIAL_SHEET)
        workbook_payload.clear()

        DataSourceFile.objects.filter(
            project=project,
            source_type='initialization',
            source_key=source_key,
        ).delete()
        source = upsert_virtual_source_file(
            project,
            'initialization',
            source_key,
            display_name,
            relative_path_value,
            {
                normalized['sheet']: {
                    'columns': normalized['columns'],
                    'rows': [],
                },
            },
        )
        sync_rows_for_model(
            project,
            source,
            InitializationWeldRow,
            normalized['sheet'],
            normalized['coreRows'],
        )
        extra_rows = normalized['extraRows']
        weld_queryset = (
            InitializationWeldRow.objects
            .filter(project=project, source_file=source, sheet_name=normalized['sheet'])
            .order_by('row_index')
        )
        extra_batch = []
        for weld, custom_fields in zip(weld_queryset.iterator(chunk_size=500), extra_rows):
            extra_batch.append(InitializationWeldExtraData(
                project=project,
                weld=weld,
                library_seq=weld.library_seq,
                custom_fields=custom_fields,
            ))
            if len(extra_batch) >= 500:
                InitializationWeldExtraData.objects.bulk_create(extra_batch, batch_size=500)
                extra_batch.clear()
        if extra_batch:
            InitializationWeldExtraData.objects.bulk_create(extra_batch, batch_size=500)

        DataSourceFile.objects.filter(
            project=project,
            source_type='idf-material',
            source_key='materials',
        ).delete()
        if material_payload is not None:
            material_source = upsert_virtual_source_file(
                project,
                'idf-material',
                'materials',
                display_name,
                relative_path_value,
                {INITIALIZATION_MATERIAL_SHEET: material_payload},
            )
            normalized_material = normalize_workbook_payload_for_models(
                {INITIALIZATION_MATERIAL_SHEET: material_payload},
                INITIALIZATION_MATERIAL_MODELS,
            )[INITIALIZATION_MATERIAL_SHEET]
            sync_rows_for_model(
                project,
                material_source,
                InitializationMaterialRow,
                INITIALIZATION_MATERIAL_SHEET,
                normalized_material['rows'],
            )
        return source


def replace_source_rows(project, source_type, source_key, display_name, relative_path_value, workbook_payload, sheet_models):
    if source_type == 'initialization':
        return replace_initialization_source_rows(
            project,
            source_key,
            display_name,
            relative_path_value,
            workbook_payload,
        )
    ensure_project_tables(project)
    with using_project_tables(project), transaction.atomic():
        DataSourceFile.objects.filter(
            project=project,
            source_type=source_type,
            source_key=source_key,
        ).delete()
    return sync_table_payload(
        project,
        source_type,
        source_key,
        display_name,
        relative_path_value,
        workbook_payload,
        sheet_models,
    )


def replace_source_sheet_rows(
    project,
    source_type,
    source_key,
    display_name,
    relative_path_value,
    sheet_name,
    sheet_payload,
    sheet_models,
):
    """Replace one worksheet while preserving every other worksheet in the source."""
    workbook_payload = {}
    source = DataSourceFile.objects.filter(
        project=project,
        source_type=source_type,
        source_key=source_key,
    ).first()
    replaced_existing_sheet = False
    if source is not None:
        for existing_sheet in source.sheet_names or []:
            if existing_sheet == sheet_name:
                workbook_payload[sheet_name] = sheet_payload
                replaced_existing_sheet = True
                continue
            selected_sheet, _, _, columns, rows = table_payload(
                source,
                sheet_models,
                existing_sheet,
            )
            workbook_payload[selected_sheet or existing_sheet] = {
                'columns': columns,
                'rows': rows,
            }
    if not replaced_existing_sheet:
        workbook_payload[sheet_name] = sheet_payload
    return replace_source_rows(
        project,
        source_type,
        source_key,
        display_name,
        relative_path_value,
        workbook_payload,
        sheet_models,
    )


def replace_source_with_workbook(project, source_type, source_key, display_name, relative_path_value, file_path, sheet_models):
    workbook_payload = read_workbook_rows(file_path)
    return replace_source_rows(
        project,
        source_type,
        source_key,
        display_name,
        relative_path_value,
        workbook_payload,
        sheet_models,
    )


def source_payload(source, sheet_models, sheet_name=None):
    if source is None:
        return '', [], 0, [], []
    return table_payload(source, sheet_models, sheet_name)


LIBRARY_MODELS = {
    'weld-library': {'*': WeldLibraryRow},
    'pipe-library': {'*': PipeMaterialRow},
    'anti-pipe-library': {'*': PipeMaterialRow},
    'fitting-library': {'*': FittingMaterialRow},
    'anti-fitting-library': {'*': FittingMaterialRow},
}

PRE_SCHEDULE_MODELS = {
    '预排产匹配结果': WeldPreScheduleRow,
    '材料匹配明细': MaterialMatchDetailRow,
}

ARRIVAL_MODELS = {
    'Sheet1': ArrivalOrderRow,
    'Sheet2': ArrivalMaterialRow,
}

INITIALIZATION_MODELS = {'*': InitializationWeldRow}

PLAN_FILE_MODELS = {
    '管段焊口表.xlsx': {'*': WeldingPlanRow},
    '下料排产单.xlsx': {'*': WeldingPlanRow},
    '防腐材料单.xlsx': {'*': AntiCorrosionMaterialOrderRow},
}

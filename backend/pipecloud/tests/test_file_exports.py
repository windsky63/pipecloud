from io import BytesIO
from unittest.mock import patch
from zipfile import ZipFile

import pandas as pd
from django.test import TestCase
from openpyxl import load_workbook

from pipecloud.models import DataSourceFile, PlanRecord, Project
from pipecloud.services.db_storage import INITIALIZATION_MODELS, replace_source_rows
from pipecloud.services.file_exports import build_project_file_tree, export_project_files
from pipecloud.services import file_export_jobs


class ProjectFileExportTests(TestCase):
    def setUp(self):
        self.project = Project.objects.create(project_name='批量导出测试')
        replace_source_rows(
            self.project,
            'initialization',
            'welds',
            '焊口初始化数据.xlsx',
            f'database://initialization/{self.project.id}/焊口初始化数据.xlsx',
            {
                'Sheet1': {
                    'columns': ['单元号', '管线号', '管段号', '焊口号', '寸径', '焊接类型', '扩展列'],
                    'rows': [{
                        '单元号': 'U1',
                        '管线号': 'L100',
                        '管段号': 'S1',
                        '焊口号': 'W-001',
                        '寸径': '10',
                        '焊接类型': 'BW',
                        '扩展列': '保留',
                    }],
                },
            },
            INITIALIZATION_MODELS,
        )

    def test_tree_and_zip_are_generated_from_project_database(self):
        tree, leaf_map = build_project_file_tree(self.project)

        self.assertEqual(tree[0]['name'], '初始化数据')
        self.assertEqual(len(leaf_map), 1)
        leaf_id = next(iter(leaf_map))

        content = export_project_files(self.project, [leaf_id])
        with ZipFile(BytesIO(content)) as archive:
            names = archive.namelist()
            self.assertEqual(names, ['初始化数据/焊口初始化数据.xlsx'])
            workbook = load_workbook(BytesIO(archive.read(names[0])), read_only=True)
            try:
                worksheet = workbook[workbook.sheetnames[0]]
                rows = worksheet.iter_rows()
                headers = [cell.value for cell in next(rows)]
                values = [cell.value for cell in next(rows)]
            finally:
                workbook.close()

        self.assertIn('扩展列', headers)
        self.assertEqual(values[headers.index('扩展列')], '保留')

    def test_export_rejects_unknown_file_id(self):
        with self.assertRaisesRegex(ValueError, '不存在或已失效'):
            export_project_files(self.project, ['source:999999'])

    def test_welding_and_cutting_tree_names_include_plan_date(self):
        for plan_key, date in [('welding', '20260701'), ('cutting', '20260702')]:
            DataSourceFile.objects.create(
                project=self.project,
                source_type='plan',
                source_key=f'{plan_key}:{date}:管段焊口表.xlsx',
                display_name='管段焊口表.xlsx',
                relative_path=f'database://plan/{plan_key}/{date}/管段焊口表.xlsx',
                sheet_names=['Sheet1'],
                sheet_columns={'Sheet1': []},
            )

        _, leaf_map = build_project_file_tree(self.project)
        names = {leaf['name'] for leaf in leaf_map.values()}

        self.assertIn('管段焊口表-2026-07-01.xlsx', names)
        self.assertIn('管段焊口表-2026-07-02.xlsx', names)

    def test_tree_does_not_generate_derived_file_rows(self):
        PlanRecord.objects.create(
            project=self.project,
            plan_key='cutting',
            plan_name='下料',
            plan_date='20260702',
            plan_folder='20260702',
            relative_path='database://plan/cutting/20260702',
            files=[{'name': '切管明细表.xlsx'}, {'name': '切管汇总表.xlsx'}],
        )

        with patch('pipecloud.services.file_exports.derived_plan_file_payload') as generate_payload:
            _, leaf_map = build_project_file_tree(self.project)

        generate_payload.assert_not_called()
        derived = [leaf for leaf in leaf_map.values() if leaf['sourceType'] == 'plan-derived']
        self.assertEqual(len(derived), 2)

    def test_selected_derived_files_share_one_bulk_generation(self):
        PlanRecord.objects.create(
            project=self.project,
            plan_key='cutting',
            plan_name='下料',
            plan_date='20260702',
            plan_folder='20260702',
            relative_path='database://plan/cutting/20260702',
            files=[{'name': '切管明细表.xlsx'}, {'name': '切管汇总表.xlsx'}],
        )
        _, leaf_map = build_project_file_tree(self.project)
        selected = [
            leaf_id
            for leaf_id, leaf in leaf_map.items()
            if leaf['sourceType'] == 'plan-derived'
        ]
        generated = {
            '切管明细表.xlsx': {'Sheet1': pd.DataFrame([{'材料代码': 'P1'}])},
            '切管汇总表.xlsx': {'Sheet1': pd.DataFrame([{'材料代码': 'P1'}])},
        }

        with patch(
            'pipecloud.services.file_exports.derived_plan_files_sheets',
            return_value=generated,
        ) as generate_files:
            content = export_project_files(self.project, selected)

        generate_files.assert_called_once()
        with ZipFile(BytesIO(content)) as archive:
            self.assertEqual(len(archive.namelist()), 2)

    def test_background_export_reports_progress_and_downloads_content(self):
        _, leaf_map = build_project_file_tree(self.project)
        selected = [next(iter(leaf_map))]

        with (
            patch.object(file_export_jobs, 'close_old_connections'),
            patch.object(
                file_export_jobs._EXECUTOR,
                'submit',
                side_effect=lambda callback, *args: callback(*args),
            ),
        ):
            job_id = file_export_jobs.start_export_job(self.project, selected)

        status = file_export_jobs.export_job_status(self.project, job_id)
        self.assertEqual(status['status'], 'completed')
        self.assertEqual(status['progress'], 100)
        content = file_export_jobs.take_export_job_content(self.project, job_id)
        with ZipFile(BytesIO(content)) as archive:
            self.assertEqual(archive.namelist(), ['初始化数据/焊口初始化数据.xlsx'])
